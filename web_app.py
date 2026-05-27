#!/usr/bin/env python3
"""
SMTP Рассылка — веб-интерфейс (GUI v2).

Локальный веб-сервер на стандартной библиотеке Python (без внешних зависимостей).
Переиспользует send_email.py (CLI) и cloud_runtime.py (облако), хранит настройки
в том же mailer_gui_config.json, что и классический интерфейс.

Запуск: python web_app.py  → автоматически откроется браузер.
"""
from __future__ import annotations

import json
import os
import re
import shutil
import subprocess
import threading
import time
import uuid
import webbrowser
from datetime import datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.error import URLError
from urllib.parse import urlparse
from urllib.request import Request, urlopen

import ssl

BASE_DIR = Path(__file__).resolve().parent
SCRIPT_PATH = BASE_DIR / "send_email.py"
CONFIG_PATH = BASE_DIR / "mailer_gui_config.json"
WEB_DIR = BASE_DIR / "web"
LOGS_DIR = BASE_DIR / "logs"
HOST = "127.0.0.1"
PORT = int(os.getenv("WEB_PORT", "8765"))

DEFAULTS: dict = {
    "template": "",
    "subject": "",
    "to_file": "",
    "email_col": "G",
    "sheet": "ALL",
    "start_row": "2",
    "fields": "A,B,C,D",
    "allow_duplicate_emails": True,
    "use_kind_template": False,
    "kind_col": "P",
    "kind_filter": "ALL",
    "extra_to": "",
    "state_file": ".send_email_state.json",
    "smtp_host": "smtp.timeweb.ru",
    "smtp_port": "465",
    "smtp_user": "",
    "smtp_password": "",
    "remember_password": True,
    "from_email": "",
    "limit_min": "20",
    "limit_day": "2000",
    "hub_url": "",
    "hub_connection_id": "",
    "hub_secret": "",
    "hub_insecure_ssl": True,
    "cloud_enabled": False,
    "server_host": "",
    "server_port": "22",
    "server_user": "",
    "server_password": "",
    "server_remote_dir": "~/mailinig-soft-cloud",
    "test_email": "",
}

# Поля, которые не отдаём в браузер открытым текстом (только флаг наличия).
SECRET_FIELDS = {"smtp_password", "hub_secret", "server_password"}


# --------------------------------------------------------------------------- #
# Конфиг
# --------------------------------------------------------------------------- #
def load_config() -> dict:
    data = dict(DEFAULTS)
    if CONFIG_PATH.exists():
        try:
            stored = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
            if isinstance(stored, dict):
                for key in DEFAULTS:
                    if key in stored:
                        data[key] = stored[key]
        except Exception:
            pass
    if not data.get("from_email"):
        data["from_email"] = data.get("smtp_user", "")
    return data


def save_config(incoming: dict) -> dict:
    current = load_config()
    for key in DEFAULTS:
        if key not in incoming:
            continue
        value = incoming[key]
        # Пустой секрет = «оставить как было».
        if key in SECRET_FIELDS and (value is None or str(value).strip() == ""):
            continue
        current[key] = value
    CONFIG_PATH.write_text(json.dumps(current, ensure_ascii=False, indent=2), encoding="utf-8")
    return current


def public_config(cfg: dict) -> dict:
    """Версия конфига для браузера: секреты заменяем на флаг наличия."""
    out = dict(cfg)
    for key in SECRET_FIELDS:
        has = bool(str(cfg.get(key, "")).strip())
        out[key] = ""
        out[f"has_{key}"] = has
    return out


def merge_secrets(settings: dict) -> dict:
    """Подставить сохранённые секреты, если браузер прислал пустые поля."""
    stored = load_config()
    merged = dict(settings)
    for key in SECRET_FIELDS:
        if not str(merged.get(key, "")).strip():
            merged[key] = stored.get(key, "")
    return merged


# --------------------------------------------------------------------------- #
# Сессии (профили) — совместимы с классическим интерфейсом
# --------------------------------------------------------------------------- #
PROFILES_PATH = BASE_DIR / "mailer_profiles.json"
PROFILE_STATE_DIR = BASE_DIR / "state_profiles"


def sanitize_filename_part(text: str) -> str:
    value = re.sub(r"[^\w\-\.]+", "_", str(text).strip(), flags=re.UNICODE).strip("._")
    return value[:80] if value else "session"


def profile_state_path(name: str) -> Path:
    safe = sanitize_filename_part(name).lower() or "profile"
    PROFILE_STATE_DIR.mkdir(parents=True, exist_ok=True)
    return PROFILE_STATE_DIR / f"{safe}.json"


def read_profiles_store() -> dict:
    if not PROFILES_PATH.exists():
        return {"active": "", "profiles": {}}
    try:
        data = json.loads(PROFILES_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {"active": "", "profiles": {}}
    if not isinstance(data, dict):
        return {"active": "", "profiles": {}}
    profiles = data.get("profiles", {})
    if not isinstance(profiles, dict):
        profiles = {}
    return {"active": str(data.get("active", "")), "profiles": profiles}


def write_profiles_store(store: dict) -> None:
    PROFILES_PATH.write_text(json.dumps(store, ensure_ascii=False, indent=2), encoding="utf-8")


def list_profiles() -> dict:
    store = read_profiles_store()
    return {"active": store.get("active", ""), "names": sorted(store.get("profiles", {}).keys())}


def save_profile(name: str, incoming: dict) -> dict:
    name = str(name).strip()
    if not name:
        raise ValueError("Укажите название сессии.")
    store = read_profiles_store()
    profiles = store.setdefault("profiles", {})
    merged = merge_secrets(incoming)
    existing = profiles.get(name) if isinstance(profiles.get(name), dict) else {}
    state_file = str(existing.get("state_file", "")).strip() or str(profile_state_path(name))
    merged["state_file"] = state_file
    profiles[name] = merged
    store["active"] = name
    write_profiles_store(store)
    save_config(merged)  # активная сессия становится текущим конфигом
    return list_profiles()


def load_profile(name: str) -> dict:
    store = read_profiles_store()
    data = store.get("profiles", {}).get(name)
    if not isinstance(data, dict):
        raise ValueError("Сессия не найдена.")
    if not str(data.get("state_file", "")).strip():
        data["state_file"] = str(profile_state_path(name))
    store["active"] = name
    write_profiles_store(store)
    save_config(data)
    return public_config(load_config())


def delete_profile(name: str) -> dict:
    store = read_profiles_store()
    profiles = store.get("profiles", {})
    profiles.pop(name, None)
    if store.get("active", "") == name:
        store["active"] = ""
    write_profiles_store(store)
    return list_profiles()


# --------------------------------------------------------------------------- #
# Сборка команды
# --------------------------------------------------------------------------- #
def _s(settings: dict, key: str) -> str:
    return str(settings.get(key, "") or "").strip()


def _b(settings: dict, key: str) -> bool:
    return bool(settings.get(key, False))


def build_argv(settings: dict, dry_run: bool, progress_file: str | None,
               override_to: list[str] | None = None, use_to_file: bool = True) -> list[str]:
    template = _s(settings, "template")
    to_file = _s(settings, "to_file")
    manual_to = [e.strip() for e in (override_to or []) if e.strip()]
    if not override_to:
        manual_to += [e.strip() for e in _s(settings, "extra_to").split(",") if e.strip()]

    if not template:
        raise ValueError("Не выбран HTML-шаблон письма.")
    if not (use_to_file and to_file) and not manual_to:
        raise ValueError("Не выбрана база получателей.")

    argv = [
        "python3", str(SCRIPT_PATH),
        "--template", template,
        "--smtp-host", _s(settings, "smtp_host") or "smtp.timeweb.ru",
        "--smtp-port", _s(settings, "smtp_port") or "465",
        "--smtp-user", _s(settings, "smtp_user"),
        "--smtp-password", _s(settings, "smtp_password"),
        "--from-email", _s(settings, "from_email") or _s(settings, "smtp_user"),
        "--xlsx-sheet", _s(settings, "sheet") or "ALL",
        "--xlsx-email-col", _s(settings, "email_col") or "G",
        "--xlsx-start-row", _s(settings, "start_row") or "2",
    ]

    if _b(settings, "use_kind_template"):
        argv += ["--xlsx-kind-col", _s(settings, "kind_col") or "P"]
    kind_filter = _s(settings, "kind_filter")
    if kind_filter:
        argv += ["--xlsx-kind-filter", kind_filter]

    hub_url = _s(settings, "hub_url")
    hub_cid = _s(settings, "hub_connection_id")
    hub_secret = _s(settings, "hub_secret")
    if hub_url and hub_cid and hub_secret:
        argv += ["--hub-url", hub_url, "--hub-connection-id", hub_cid, "--hub-module-secret", hub_secret]
        if _b(settings, "hub_insecure_ssl"):
            argv += ["--hub-insecure-ssl"]

    if _b(settings, "allow_duplicate_emails"):
        argv += ["--allow-duplicate-emails"]

    subject = _s(settings, "subject")
    if subject:
        argv += ["--subject", subject]
    fields = _s(settings, "fields")
    if fields:
        argv += ["--xlsx-fields", fields]
    if _s(settings, "limit_min"):
        argv += ["--limit-per-minute", _s(settings, "limit_min")]
    if _s(settings, "limit_day"):
        argv += ["--limit-per-day", _s(settings, "limit_day")]

    if use_to_file and to_file:
        argv += ["--to-file", to_file]
    if _s(settings, "state_file"):
        argv += ["--state-file", _s(settings, "state_file")]

    for email in manual_to:
        argv += ["--to", email]

    if progress_file:
        argv += ["--progress-file", progress_file]
    if dry_run:
        argv += ["--dry-run"]
    return argv


def build_remote_argv(argv: list[str], remote_base: str) -> list[str]:
    remote_base = remote_base.rstrip("/")
    path_flags = {"--template", "--to-file", "--state-file", "--progress-file"}

    def map_path(raw: str) -> str:
        path = Path(raw).expanduser()
        try:
            resolved = path.resolve()
        except Exception:
            return raw
        if resolved == SCRIPT_PATH:
            return f"{remote_base}/send_email.py"
        try:
            if resolved.is_relative_to(BASE_DIR):
                return f"{remote_base}/{resolved.relative_to(BASE_DIR).as_posix()}"
        except Exception:
            pass
        return raw

    out: list[str] = []
    prev = ""
    for i, token in enumerate(argv):
        if i == 0:
            out.append("python3")
            prev = ""
            continue
        if i == 1:
            out.append(map_path(token))
            prev = ""
            continue
        if token.startswith("--"):
            out.append(token)
            prev = token
            continue
        out.append(map_path(token) if prev in path_flags else token)
        prev = ""
    return out


# --------------------------------------------------------------------------- #
# Управление запуском
# --------------------------------------------------------------------------- #
class RunManager:
    def __init__(self) -> None:
        self.lock = threading.Lock()
        self.execution = "local"          # local | cloud
        self.process: subprocess.Popen | None = None
        self.progress_path: Path | None = None
        self.log_path: Path | None = None
        self.dry_run = False
        # cloud
        self.cloud = None                 # CloudRuntime
        self.cloud_run: dict | None = None
        self.cloud_remote_progress = ""
        self.cloud_log_offset = 0
        self.cloud_log_buffer = ""
        self.last_error = ""
        self.started_at = ""

    def is_running(self) -> bool:
        if self.execution == "local":
            return self.process is not None and self.process.poll() is None
        return self.cloud_run is not None

    def _new_paths(self, tag: str) -> tuple[Path, Path]:
        LOGS_DIR.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        base = LOGS_DIR / f"{stamp}_{tag}"
        return base.with_suffix(".log"), base.with_suffix(".progress.json")

    def start(self, settings: dict, dry_run: bool, execution: str,
              override_to: list[str] | None = None, use_to_file: bool = True) -> None:
        with self.lock:
            if self.is_running():
                raise RuntimeError("Уже выполняется другой запуск. Остановите его сначала.")
            self.last_error = ""
            self.dry_run = dry_run
            self.execution = execution
            self.started_at = datetime.now().strftime("%H:%M:%S")
            tag = ("test" if override_to else "send") + ("_dry" if dry_run else "")
            if execution == "cloud":
                self._start_cloud(settings, dry_run, tag, override_to, use_to_file)
            else:
                self._start_local(settings, dry_run, tag, override_to, use_to_file)

    def _start_local(self, settings, dry_run, tag, override_to, use_to_file) -> None:
        log_path, progress_path = self._new_paths(tag)
        argv = build_argv(settings, dry_run, str(progress_path), override_to, use_to_file)
        self.log_path = log_path
        self.progress_path = progress_path
        log_handle = open(log_path, "w", encoding="utf-8")
        log_handle.write("Команда:\n" + sanitize_cmd(argv) + "\n\n")
        log_handle.flush()
        self.process = subprocess.Popen(
            argv, stdout=log_handle, stderr=subprocess.STDOUT, cwd=str(BASE_DIR),
        )

    def _start_cloud(self, settings, dry_run, tag, override_to, use_to_file) -> None:
        from cloud_runtime import CloudRuntime, ServerConfig

        remote_dir = _s(settings, "server_remote_dir") or "~/mailinig-soft-cloud"
        config = ServerConfig(
            host=_s(settings, "server_host"),
            port=int(_s(settings, "server_port") or "22"),
            username=_s(settings, "server_user"),
            password=_s(settings, "server_password"),
            remote_dir=remote_dir,
        )
        if not config.host or not config.username:
            raise RuntimeError("Для облака заполните адрес сервера и логин на вкладке «Сервер».")

        stamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        remote_progress = f"{remote_dir.rstrip('/')}/logs/{stamp}_{tag}.progress.json"
        argv = build_argv(settings, dry_run, remote_progress, override_to, use_to_file)
        remote_argv = build_remote_argv(argv, remote_dir)

        runtime = CloudRuntime(config, BASE_DIR)
        runtime.connect()
        runtime.upload_project()
        run_info = runtime.start_remote_process_detached(remote_argv)
        self.cloud = runtime
        self.cloud_run = run_info
        self.cloud_remote_progress = remote_progress
        self.cloud_log_offset = 0
        self.cloud_log_buffer = ""

    def progress(self) -> dict:
        if self.execution == "cloud":
            return self._cloud_progress()
        return self._local_progress()

    def _read_progress_file(self, path: Path | None) -> dict:
        if not path or not path.exists():
            return {}
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            return {}

    def _local_progress(self) -> dict:
        running = self.is_running()
        prog = self._read_progress_file(self.progress_path)
        log_text = ""
        if self.log_path and self.log_path.exists():
            try:
                log_text = self.log_path.read_text(encoding="utf-8", errors="replace")
            except Exception:
                log_text = ""
        finished = self.process is not None and self.process.poll() is not None
        return {
            "execution": "local",
            "running": running,
            "finished": finished and not running,
            "dry_run": self.dry_run,
            "started_at": self.started_at,
            "progress": prog,
            "log": log_text[-12000:],
            "error": self.last_error,
        }

    def _cloud_progress(self) -> dict:
        if self.cloud is None or self.cloud_run is None:
            return {"execution": "cloud", "running": False, "finished": True, "progress": {}, "log": "", "error": self.last_error}
        prog = {}
        try:
            text = self.cloud.download_text_file(self.cloud_remote_progress)
            prog = json.loads(text)
        except Exception:
            prog = {}
        try:
            chunk, self.cloud_log_offset = self.cloud.read_log_chunk(
                self.cloud_run["log_file"], self.cloud_log_offset
            )
            if chunk:
                self.cloud_log_buffer += chunk
        except Exception:
            pass
        try:
            running = self.cloud.is_remote_process_running(self.cloud_run["pid_file"])
        except Exception:
            running = False
        return {
            "execution": "cloud",
            "running": running,
            "finished": not running,
            "dry_run": self.dry_run,
            "started_at": self.started_at,
            "progress": prog,
            "log": self.cloud_log_buffer[-12000:],
            "error": self.last_error,
        }

    def stop(self) -> str:
        with self.lock:
            if self.execution == "cloud" and self.cloud and self.cloud_run:
                try:
                    ok, message = self.cloud.stop_remote_process(self.cloud_run["run_id"])
                    return message
                except Exception as error:
                    return f"Ошибка остановки: {error}"
            if self.process and self.process.poll() is None:
                self.process.terminate()
                return "Остановка запрошена."
            return "Активного запуска нет."

    def reset_cloud(self) -> None:
        with self.lock:
            if self.cloud is not None:
                try:
                    self.cloud.close()
                except Exception:
                    pass
            self.cloud = None
            self.cloud_run = None


MANAGER = RunManager()


def sanitize_cmd(argv: list[str]) -> str:
    safe, mask = [], False
    for token in argv:
        if mask:
            safe.append("********")
            mask = False
            continue
        safe.append(token)
        if token in {"--smtp-password", "--smtp-account"}:
            mask = True
    return " ".join(safe)


# --------------------------------------------------------------------------- #
# Вспомогательные действия
# --------------------------------------------------------------------------- #
def native_file_dialog(kind: str) -> str:
    """Открыть нативный диалог macOS и вернуть POSIX-путь выбранного файла."""
    if kind == "excel":
        prompt, types = "Выберите файл базы (Excel/CSV)", '{"xlsx","csv","txt"}'
    else:
        prompt, types = "Выберите HTML-письмо", '{"html","htm"}'
    script = f'POSIX path of (choose file with prompt "{prompt}" of type {types})'
    try:
        out = subprocess.run(["osascript", "-e", script], capture_output=True, text=True, timeout=300)
        if out.returncode == 0:
            return out.stdout.strip()
    except Exception:
        pass
    # Fallback без фильтра типов (на случай отказа фильтра).
    try:
        out = subprocess.run(
            ["osascript", "-e", f'POSIX path of (choose file with prompt "{prompt}")'],
            capture_output=True, text=True, timeout=300,
        )
        if out.returncode == 0:
            return out.stdout.strip()
    except Exception:
        pass
    return ""


def detect_email_columns(to_file: str) -> list[str]:
    import send_email  # модуль лежит рядом; обработчики сигналов ставятся только в main()
    return send_email.detect_xlsx_email_columns(Path(to_file).expanduser().resolve(), "ALL")


def analyze_workbook(to_file: str) -> dict:
    """Умный разбор: сам находит колонку email, строку начала данных,
    считает получателей и разбирается с несколькими вкладками."""
    import send_email
    path = Path(to_file).expanduser().resolve()
    if not path.exists():
        return {"ok": False, "message": "Файл не найден."}

    ext = path.suffix.lower()
    if ext in {".csv", ".txt"}:
        content = path.read_text(encoding="utf-8", errors="ignore")
        emails = {m.lower() for m in send_email.EMAIL_RE.findall(content)}
        if not emails:
            return {"ok": False, "message": "В файле не нашлось ни одного email."}
        return {"ok": True, "email_col": "A", "start_row": 1, "total": len(emails),
                "sheets": 1, "columns": ["A"],
                "message": f"Найдено {len(emails)} адресов в текстовом файле."}

    if ext != ".xlsx":
        return {"ok": False, "message": "Поддерживаются .xlsx, .csv, .txt."}

    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        return {"ok": False, "message": "Не установлен openpyxl."}

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        col_hits: dict[int, int] = {}              # сколько email в каждой колонке (по всем листам)
        col_first_row: dict[int, int] = {}         # первая строка с email в колонке
        unique_emails: dict[int, set] = {}         # уникальные адреса по колонке
        sheet_count = 0
        for sheet in wb.worksheets:
            sheet_count += 1
            for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=5000, values_only=True), start=1):
                for col_idx, value in enumerate(row, start=1):
                    text = send_email.to_str(value)
                    if text and send_email.EMAIL_RE.fullmatch(text):
                        col_hits[col_idx] = col_hits.get(col_idx, 0) + 1
                        if col_idx not in col_first_row:
                            col_first_row[col_idx] = row_idx
                        unique_emails.setdefault(col_idx, set()).add(text.lower())
        if not col_hits:
            return {"ok": False, "message": "Не нашёл ни одного email ни на одной вкладке. "
                                            "Проверь, что в файле есть адреса."}
        # Лучшая колонка — где больше всего адресов
        best_col = max(col_hits.items(), key=lambda kv: kv[1])[0]
        best_letter = get_column_letter(best_col)
        start_row = col_first_row.get(best_col, 1)
        total = len(unique_emails.get(best_col, set()))
        ranked = [get_column_letter(c) for c, _ in sorted(col_hits.items(), key=lambda kv: kv[1], reverse=True)[:5]]
        sheets_note = f", вкладок: {sheet_count} (объединяю все)" if sheet_count > 1 else ""
        return {
            "ok": True,
            "email_col": best_letter,
            "start_row": start_row,
            "total": total,
            "sheets": sheet_count,
            "columns": ranked,
            "message": f"Нашёл {total} получателей · колонка {best_letter} · с {start_row}-й строки{sheets_note}.",
        }
    finally:
        wb.close()


def hub_check(settings: dict) -> dict:
    url = _s(settings, "hub_url")
    if not url:
        return {"ok": False, "message": "Не указан адрес Хаба."}
    ctx = ssl._create_unverified_context() if _b(settings, "hub_insecure_ssl") else None
    try:
        with urlopen(Request(url.rstrip("/") + "/index.php", method="GET"), timeout=15, context=ctx) as resp:
            return {"ok": True, "message": f"Хаб доступен (HTTP {resp.getcode()})."}
    except URLError as error:
        return {"ok": False, "message": f"Хаб недоступен: {error}"}


def cloud_check(settings: dict) -> dict:
    from cloud_runtime import CloudRuntime, ServerConfig
    config = ServerConfig(
        host=_s(settings, "server_host"),
        port=int(_s(settings, "server_port") or "22"),
        username=_s(settings, "server_user"),
        password=_s(settings, "server_password"),
        remote_dir=_s(settings, "server_remote_dir") or "~/mailinig-soft-cloud",
    )
    if not config.host or not config.username:
        return {"ok": False, "message": "Заполните адрес сервера и логин."}
    try:
        runtime = CloudRuntime(config, BASE_DIR)
        runtime.connect()
        runtime.close()
        return {"ok": True, "message": f"Подключение к {config.host} успешно."}
    except Exception as error:
        message = str(error)
        if "publickey" in message:
            message = ("Сервер пускает только по SSH-ключу, а не по паролю. "
                       "Включите вход по паролю на сервере или добавьте ключ.")
        return {"ok": False, "message": f"Не удалось подключиться: {message}"}


def _resolve_state_path(state_file: str) -> Path:
    raw = (state_file or "").strip() or ".send_email_state.json"
    path = Path(raw).expanduser()
    if not path.is_absolute():
        path = BASE_DIR / path
    return path


def reset_daily_counter(state_file: str) -> dict:
    """Обнулить суточный счётчик отправки (как кнопка в классике)."""
    from datetime import date
    path = _resolve_state_path(state_file)
    data = {}
    if path.exists():
        try:
            loaded = json.loads(path.read_text(encoding="utf-8"))
            if isinstance(loaded, dict):
                data = loaded
        except Exception:
            data = {}
    data["date"] = date.today().isoformat()
    data["sent_today"] = 0
    data["account_sent_today"] = {}
    if not isinstance(data.get("campaigns"), dict):
        data["campaigns"] = {}
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"ok": True, "message": "Суточный счётчик обнулён — можно отправлять дальше."}


def reset_campaign_progress(state_file: str) -> dict:
    """Сбросить прогресс кампаний в state-файле (рассылка начнётся с начала базы)."""
    path = _resolve_state_path(state_file)
    if not path.exists():
        return {"ok": True, "message": "Прогресс уже пуст — рассылка начнётся с начала."}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        data = {}
    if not isinstance(data, dict):
        data = {}
    data["campaigns"] = {}
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"ok": True, "message": "Прогресс сброшен — следующий запуск начнётся с начала базы."}


def count_recipients(settings: dict) -> dict:
    """Быстрый dry-run для подсчёта получателей и проверки колонки/шаблона."""
    merged = merge_secrets(settings)
    try:
        argv = build_argv(merged, dry_run=True, progress_file=None)
    except ValueError as error:
        return {"ok": False, "message": str(error)}
    try:
        out = subprocess.run(argv, capture_output=True, text=True, timeout=300, cwd=str(BASE_DIR))
    except Exception as error:
        return {"ok": False, "message": f"Ошибка проверки: {error}"}
    text = (out.stdout or "") + (out.stderr or "")
    if out.returncode != 0:
        first = next((ln for ln in text.splitlines() if ln.strip()), "Ошибка проверки.")
        return {"ok": False, "message": first.strip(), "log": text[-4000:]}
    total = None
    for line in text.splitlines():
        if "осталось" in line:
            digits = "".join(ch for ch in line if ch.isdigit())
            if digits:
                total = int(digits)
                break
    return {"ok": True, "total": total, "log": text[-4000:]}


# --------------------------------------------------------------------------- #
# HTTP
# --------------------------------------------------------------------------- #
class Handler(BaseHTTPRequestHandler):
    def log_message(self, *_args) -> None:  # тише в консоли
        pass

    def _send_json(self, data: dict, status: int = 200) -> None:
        body = json.dumps(data, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _send_file(self, path: Path, content_type: str) -> None:
        if not path.exists():
            self.send_error(404, "Not found")
            return
        body = path.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _read_body(self) -> dict:
        length = int(self.headers.get("Content-Length", "0") or "0")
        if length <= 0:
            return {}
        raw = self.rfile.read(length)
        try:
            return json.loads(raw.decode("utf-8"))
        except Exception:
            return {}

    def do_GET(self) -> None:
        route = urlparse(self.path).path
        if route in ("/", "/index.html"):
            self._send_file(WEB_DIR / "index.html", "text/html; charset=utf-8")
        elif route == "/app.js":
            self._send_file(WEB_DIR / "app.js", "application/javascript; charset=utf-8")
        elif route == "/styles.css":
            self._send_file(WEB_DIR / "styles.css", "text/css; charset=utf-8")
        elif route == "/api/config":
            self._send_json(public_config(load_config()))
        elif route == "/api/profiles":
            self._send_json(list_profiles())
        elif route == "/api/progress":
            self._send_json(MANAGER.progress())
        else:
            self.send_error(404, "Not found")

    def do_POST(self) -> None:
        route = urlparse(self.path).path
        body = self._read_body()
        try:
            if route == "/api/config":
                save_config(body.get("settings", {}))
                self._send_json({"ok": True})
            elif route == "/api/browse":
                path = native_file_dialog(body.get("kind", "excel"))
                self._send_json({"ok": bool(path), "path": path})
            elif route == "/api/detect-column":
                cols = detect_email_columns(body.get("to_file", ""))
                self._send_json({"ok": bool(cols), "columns": cols})
            elif route == "/api/analyze":
                self._send_json(analyze_workbook(body.get("to_file", "")))
            elif route == "/api/profiles/save":
                self._send_json({"ok": True, **save_profile(body.get("name", ""), body.get("settings", {}))})
            elif route == "/api/profiles/load":
                self._send_json({"ok": True, "settings": load_profile(body.get("name", ""))})
            elif route == "/api/profiles/delete":
                self._send_json({"ok": True, **delete_profile(body.get("name", ""))})
            elif route == "/api/count":
                self._send_json(count_recipients(body.get("settings", {})))
            elif route == "/api/reset-daily":
                self._send_json(reset_daily_counter(body.get("state_file", "")))
            elif route == "/api/reset-progress":
                self._send_json(reset_campaign_progress(body.get("state_file", "")))
            elif route == "/api/hub-check":
                self._send_json(hub_check(merge_secrets(body.get("settings", {}))))
            elif route == "/api/cloud-check":
                self._send_json(cloud_check(merge_secrets(body.get("settings", {}))))
            elif route == "/api/start":
                settings = merge_secrets(body.get("settings", {}))
                save_config(body.get("settings", {}))
                dry_run = bool(body.get("dry_run", False))
                execution = "cloud" if settings.get("cloud_enabled") else "local"
                override_to = body.get("override_to") or None
                use_to_file = not override_to
                MANAGER.start(settings, dry_run, execution, override_to, use_to_file)
                self._send_json({"ok": True, "execution": execution})
            elif route == "/api/stop":
                self._send_json({"ok": True, "message": MANAGER.stop()})
            elif route == "/api/cloud-reset":
                MANAGER.reset_cloud()
                self._send_json({"ok": True})
            else:
                self.send_error(404, "Not found")
        except Exception as error:
            self._send_json({"ok": False, "message": str(error)}, status=400)


def main() -> None:
    WEB_DIR.mkdir(exist_ok=True)
    server = ThreadingHTTPServer((HOST, PORT), Handler)
    url = f"http://{HOST}:{PORT}/"
    print(f"SMTP Рассылка (веб) запущена: {url}")
    print("Чтобы закрыть — закройте это окно терминала.")
    threading.Thread(target=lambda: (time.sleep(0.8), webbrowser.open(url)), daemon=True).start()
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        server.server_close()


if __name__ == "__main__":
    main()
