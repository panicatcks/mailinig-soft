#!/usr/bin/env python3
"""Проверка email-адресов: синтаксис + наличие MX/A записи у домена.

DNS проверяется через nslookup, а при недоступности системного DNS —
через DNS-over-HTTPS (Cloudflare/Google) с проверкой TLS.

CLI:
    python email_validator.py --in база.xlsx --email-col G --out очищенная.xlsx --bad bad.txt
    python email_validator.py --in base.txt --out clean.txt --bad bad.txt

Модульное API:
    from email_validator import validate_emails
    results = validate_emails(["a@b.ru", "x@y.zz"], on_progress=print)
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import ssl
import subprocess
import sys
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Callable, Iterable
from urllib.parse import urlencode
from urllib.request import Request, urlopen

EMAIL_RE = re.compile(
    r"^[A-Za-z0-9.!#$%&'*+/=?^_`{|}~\-]+@[A-Za-z0-9.\-]+$"
)
DOMAIN_RE = re.compile(r"^[A-Za-z0-9.\-]+$")

_dns_cache: dict[str, tuple[bool | None, str]] = {}
_dns_lock = threading.Lock()
_system_dns_available: bool | None = None


def email_syntax_reason(email: object) -> str:
    """Return ``ok`` or a stable reason for rejecting an address.

    Deliberately accepts the common, unquoted form only.  A mailing database is
    safer when an exotic-but-legal address is reported for manual review than
    when a broken substring is silently sent to.
    """
    value = "" if email is None else str(email).strip()
    if not value:
        return "empty"
    if len(value) > 254:
        return "too_long"
    if any(char.isspace() for char in value):
        return "whitespace"
    if value.count("@") != 1 or not EMAIL_RE.fullmatch(value):
        return "bad_syntax"

    local, domain = value.rsplit("@", 1)
    if len(local) > 64:
        return "local_too_long"
    if local.startswith(".") or local.endswith(".") or ".." in local:
        return "bad_local_part"

    if domain.endswith("."):
        return "bad_domain"
    domain = domain.lower()
    if not domain or len(domain) > 253 or not DOMAIN_RE.fullmatch(domain):
        return "bad_domain"
    labels = domain.split(".")
    if len(labels) < 2:
        return "bad_domain"
    for label in labels:
        if not label or len(label) > 63:
            return "bad_domain"
        if not label[0].isalnum() or not label[-1].isalnum():
            return "bad_domain"
    tld = labels[-1]
    if not (tld.isalpha() and len(tld) >= 2) and not tld.startswith("xn--"):
        return "bad_domain"
    return "ok"


def is_valid_email_syntax(email: object) -> bool:
    return email_syntax_reason(email) == "ok"


def _nslookup(domain: str, record: str) -> tuple[str, bool, bool]:
    """Return ``(output, dns_alive, query_succeeded)``.

    ``nslookup`` prints the DNS server's own ``Address:`` even for NXDOMAIN.
    Keeping the process result separate prevents that line from being mistaken
    for an A record (the old implementation accepted almost every fake domain).
    """
    global _system_dns_available
    with _dns_lock:
        if _system_dns_available is False:
            return "", False, False
    try:
        proc = subprocess.run(
            ["nslookup", f"-type={record}", domain],
            capture_output=True,
            text=True,
            timeout=6,
        )
        out = (proc.stdout or "") + (proc.stderr or "")
        dead_markers = ("connection timed out", "no servers could be reached",
                        "command not found", "operation timed out")
        lowered = out.lower()
        alive = not any(marker in lowered for marker in dead_markers)
        with _dns_lock:
            _system_dns_available = alive
        negative_markers = (
            "nxdomain", "non-existent domain", "server can't find",
            "no answer", "not found",
        )
        succeeded = proc.returncode == 0 and not any(
            marker in lowered for marker in negative_markers
        )
        return out, alive, succeeded
    except FileNotFoundError:
        with _dns_lock:
            _system_dns_available = False
        return "", False, False
    except subprocess.TimeoutExpired:
        with _dns_lock:
            _system_dns_available = False
        return "", False, False
    except Exception:
        with _dns_lock:
            _system_dns_available = False
        return "", False, False


def _doh_lookup(domain: str, record: str) -> list[str] | None:
    """Resolve through DNS-over-HTTPS; ``None`` means all providers failed."""
    try:
        import certifi

        context = ssl.create_default_context(cafile=certifi.where())
    except ImportError:  # pragma: no cover - depends on installation
        context = ssl.create_default_context()

    providers = (
        "https://cloudflare-dns.com/dns-query",
        "https://dns.google/resolve",
    )
    for base_url in providers:
        url = base_url + "?" + urlencode({"name": domain, "type": record})
        request = Request(
            url,
            headers={"Accept": "application/dns-json", "User-Agent": "mailer-validator/1"},
        )
        try:
            with urlopen(request, timeout=7, context=context) as response:
                payload = json.loads(response.read().decode("utf-8", errors="replace"))
            status = int(payload.get("Status", -1))
            if status not in (0, 3):
                continue
            answers = payload.get("Answer", [])
            if not isinstance(answers, list):
                answers = []
            wanted_type = 15 if record.upper() == "MX" else 1
            return [
                str(item.get("data", "")).strip()
                for item in answers
                if isinstance(item, dict) and int(item.get("type", -1)) == wanted_type
            ]
        except Exception:
            continue
    return None


def domain_mail_status(domain: str) -> tuple[bool | None, str]:
    """Check MX/implicit MX. ``None`` means DNS itself was unavailable."""
    domain = domain.strip().lower().rstrip(".")
    if email_syntax_reason(f"x@{domain}") != "ok":
        return False, "bad_domain"
    with _dns_lock:
        if domain in _dns_cache:
            return _dns_cache[domain]

    mx_out, mx_alive, mx_ok = _nslookup(domain, "MX")
    mx_lower = mx_out.lower()
    # RFC 7505 null MX: the domain explicitly accepts no mail.
    if mx_ok and re.search(r"mail exchanger\s*=\s*\.\s*$", mx_out, re.MULTILINE | re.IGNORECASE):
        result: tuple[bool | None, str] = (False, "null_mx")
    elif mx_ok and ("mail exchanger" in mx_lower or "mx preference" in mx_lower):
        result = (True, "mx")
    else:
        a_out, a_alive, a_ok = _nslookup(domain, "A")
        # Only inspect the answer section after Name:, never the resolver's own
        # Address line printed at the top of nslookup output.
        name_match = re.search(r"^Name:\s*.+$", a_out, re.MULTILINE | re.IGNORECASE)
        answer_tail = a_out[name_match.end():] if name_match else ""
        has_address = bool(re.search(r"^Addresses?:\s*\S+", answer_tail, re.MULTILINE | re.IGNORECASE))
        if a_ok and name_match and has_address:
            result = (True, "implicit_mx")
        elif mx_alive and a_alive:
            result = (False, "no_mx")
        else:
            doh_mx = _doh_lookup(domain, "MX")
            if doh_mx is None:
                result = (None, "dns_unavailable")
            elif any(value in {".", "0 ."} for value in doh_mx):
                result = (False, "null_mx")
            elif doh_mx:
                result = (True, "mx")
            else:
                doh_a = _doh_lookup(domain, "A")
                if doh_a is None:
                    result = (None, "dns_unavailable")
                elif doh_a:
                    result = (True, "implicit_mx")
                else:
                    result = (False, "no_mx")

    if result[0] is not None:
        with _dns_lock:
            _dns_cache[domain] = result
    return result


def domain_has_mail(domain: str) -> bool:
    """True если у домена есть MX или хотя бы A запись (RFC 5321 implicit MX)."""
    status, _reason = domain_mail_status(domain)
    # Preserve the public bool API: an unavailable resolver is not proof that a
    # real address is bad, so callers that only ask for bool must not delete it.
    return status is not False


def validate_one(email: str) -> dict:
    """Возвращает {'email': ..., 'ok': bool, 'reason': ...}."""
    e = (email or "").strip()
    syntax_reason = email_syntax_reason(e)
    if syntax_reason != "ok":
        return {"email": email, "ok": False, "reason": syntax_reason, "verified": False}
    domain = e.rsplit("@", 1)[-1]
    status, reason = domain_mail_status(domain)
    if status is False:
        return {"email": email, "ok": False, "reason": reason, "verified": True}
    return {
        "email": email,
        "ok": True,
        "reason": "ok" if status else reason,
        "verified": status is not None,
    }


def validate_emails(
    emails: Iterable[str],
    workers: int = 16,
    on_progress: Callable[[int, int, dict], None] | None = None,
) -> list[dict]:
    """Параллельная проверка списка. on_progress(done, total, result)."""
    items = list(emails)
    total = len(items)
    results: list[dict | None] = [None] * total
    done = 0
    lock = threading.Lock()

    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = {pool.submit(validate_one, e): i for i, e in enumerate(items)}
        for fut in as_completed(futures):
            i = futures[fut]
            try:
                results[i] = fut.result()
            except Exception as exc:
                results[i] = {"email": items[i], "ok": False, "reason": f"error: {exc}"}
            with lock:
                done += 1
                if on_progress:
                    on_progress(done, total, results[i])
    return [r for r in results if r is not None]


# ---------- I/O helpers ----------

def _read_text_emails(path: Path) -> list[str]:
    text = path.read_text(encoding="utf-8", errors="ignore")
    return [line.strip() for line in text.splitlines() if line.strip()]


def _col_to_index(col: str) -> int:
    col = (col or "A").strip().upper()
    idx = 0
    for ch in col:
        if not ch.isalpha():
            break
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return max(idx - 1, 0)


def _read_xlsx_emails(path: Path, email_col: str, start_row: int = 2) -> tuple[list[str], list[int]]:
    """Возвращает (emails, row_numbers) — параллельные списки."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("openpyxl не установлен. Используй .txt/.csv базу или поставь openpyxl.")
    wb = load_workbook(path, read_only=True, data_only=True)
    col_idx = _col_to_index(email_col)
    emails: list[str] = []
    rows: list[int] = []
    for ws in wb.worksheets:
        for r, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if r < start_row:
                continue
            # Ignore truly blank worksheet rows, but include an empty/garbage
            # email cell when the row contains other customer data.  Previously
            # those rows never reached validation and survived every cleanup.
            if not any(value is not None and str(value).strip() for value in row):
                continue
            value = row[col_idx] if col_idx < len(row) else None
            text = "" if value is None else str(value).strip()
            emails.append(text)
            rows.append(r)
    wb.close()
    return emails, rows


def _write_xlsx_cleaned(src: Path, dest: Path, bad_set: set[str], email_col: str, start_row: int = 2) -> int:
    """Перезаписывает xlsx, удаляя строки с плохими адресами. Возвращает кол-во удалённых."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("openpyxl нужен для записи xlsx.")
    wb = load_workbook(src)
    col_idx = _col_to_index(email_col)
    removed = 0
    bad_keys = {str(value).strip().lower() for value in bad_set}
    for ws in wb.worksheets:
        rows_to_delete = []
        for r in range(start_row, ws.max_row + 1):
            cell = ws.cell(row=r, column=col_idx + 1).value
            row_has_data = any(
                value is not None and str(value).strip()
                for value in next(ws.iter_rows(min_row=r, max_row=r, values_only=True))
            )
            if not row_has_data:
                continue
            email = "" if cell is None else str(cell).strip()
            if email.lower() in bad_keys or not is_valid_email_syntax(email):
                rows_to_delete.append(r)
        for r in reversed(rows_to_delete):
            ws.delete_rows(r, 1)
            removed += 1
    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)
    return removed


def _write_xlsx_clean_dedup(
    src: Path,
    dest: Path,
    bad_set: set[str],
    dedup: bool,
    email_col: str,
    start_row: int = 2,
    drop_bad_syntax: bool = False,
) -> tuple[int, int, list[str]]:
    """Чистит xlsx: удаляет невалидные и (опц.) повторные адреса по всем листам.

    Дубликаты считаются глобально по всей книге (первое вхождение остаётся).
    drop_bad_syntax: дополнительно удалять ячейки без корректного email-синтаксиса
    (в т.ч. значения без «@» — их не видит _read_xlsx_emails).
    Возвращает (удалено_невалидных, удалено_дубликатов, список_удалённых_значений).
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("openpyxl нужен для записи xlsx.")
    wb = load_workbook(src)
    col_idx = _col_to_index(email_col)
    removed_bad = 0
    removed_dup = 0
    removed_values: list[str] = []
    seen: set[str] = set()
    bad_keys = {str(value).strip().lower() for value in bad_set}
    for ws in wb.worksheets:
        rows_to_delete: list[int] = []
        for r in range(start_row, ws.max_row + 1):
            cell = ws.cell(row=r, column=col_idx + 1).value
            row_has_data = any(
                value is not None and str(value).strip()
                for value in next(ws.iter_rows(min_row=r, max_row=r, values_only=True))
            )
            if not row_has_data:
                continue
            email = "" if cell is None else str(cell).strip()
            key = email.lower()
            if key in bad_keys or (drop_bad_syntax and not is_valid_email_syntax(email)):
                rows_to_delete.append(r)
                removed_bad += 1
                removed_values.append(email or "<empty>")
                continue
            if dedup:
                if key in seen:
                    rows_to_delete.append(r)
                    removed_dup += 1
                    removed_values.append(email)
                    continue
                seen.add(key)
        for r in reversed(rows_to_delete):
            ws.delete_rows(r, 1)
    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)
    return removed_bad, removed_dup, removed_values


def main() -> None:
    ap = argparse.ArgumentParser(description="Очистка списка email от невалидных адресов.")
    ap.add_argument("--in", dest="src", required=True, help="Исходный файл (.xlsx/.csv/.txt)")
    ap.add_argument("--email-col", default="A", help="Колонка email в xlsx (по умолч. A)")
    ap.add_argument("--start-row", type=int, default=2, help="С какой строки читать xlsx (по умолч. 2)")
    ap.add_argument("--out", default="", help="Куда сохранить очищенный файл (необязательно)")
    ap.add_argument("--txt-out", dest="txt_out", default="", help="Куда сохранить email через запятую (.txt)")
    ap.add_argument("--bad", default="", help="Куда выписать невалидные адреса (txt)")
    ap.add_argument("--report", default="", help="CSV-отчёт со статусом по каждому адресу")
    ap.add_argument("--dedup", action="store_true", help="Удалять дубликаты email")
    ap.add_argument("--skip-mx", dest="skip_mx", action="store_true",
                    help="Не проверять MX/домен — только синтаксис и дубликаты (быстро, без интернета)")
    ap.add_argument("--workers", type=int, default=16)
    args = ap.parse_args()

    src = Path(args.src).expanduser().resolve()
    if not src.exists():
        sys.exit(f"Файл не найден: {src}")

    is_xlsx = src.suffix.lower() in {".xlsx", ".xlsm"}
    if is_xlsx:
        emails, _rows = _read_xlsx_emails(src, args.email_col, args.start_row)
    else:
        emails = _read_text_emails(src)

    if not emails:
        sys.exit("В файле нет адресов.")

    total = len(emails)

    if args.skip_mx:
        print(f"Загружено {total} адресов. Проверяю только синтаксис…", flush=True)
        results = [
            {"email": e, "ok": is_valid_email_syntax(e),
             "reason": email_syntax_reason(e)}
            for e in emails
        ]
    else:
        print(f"Загружено {total} адресов. Проверяю MX/синтаксис…", flush=True)
        last_pct = -1

        def progress(done: int, total: int, r: dict) -> None:
            nonlocal last_pct
            pct = int(done / total * 100)
            if pct != last_pct and pct % 5 == 0:
                print(f"  {done}/{total} ({pct}%)", flush=True)
                last_pct = pct

        results = validate_emails(emails, workers=args.workers, on_progress=progress)

    bad = [r for r in results if not r["ok"]]
    ok = [r for r in results if r["ok"]]
    bad_set = {r["email"].strip() for r in bad}

    print(f"\nГотово. Валидных: {len(ok)}, плохих: {len(bad)}")
    unverified = sum(1 for result in ok if not result.get("verified", True))
    if unverified:
        print(f"  не удалось проверить DNS: {unverified} (адреса не удалены)")
    by_reason: dict[str, int] = {}
    for r in bad:
        by_reason[r["reason"]] = by_reason.get(r["reason"], 0) + 1
    for reason, count in sorted(by_reason.items(), key=lambda x: -x[1]):
        print(f"  {reason}: {count}")

    def dedup_keep_order(items: list[str]) -> tuple[list[str], int]:
        seen: set[str] = set()
        kept: list[str] = []
        removed = 0
        for e in items:
            key = e.strip().lower()
            if not key:
                continue
            if key in seen:
                removed += 1
                continue
            seen.add(key)
            kept.append(e)
        return kept, removed

    if args.bad:
        bad_path = Path(args.bad).expanduser().resolve()
        bad_path.parent.mkdir(parents=True, exist_ok=True)
        bad_path.write_text(
            "\n".join(str(r["email"]).strip() or "<empty>" for r in bad),
            encoding="utf-8",
        )
        print(f"Плохие выписаны в {bad_path}")

    if args.report:
        rep_path = Path(args.report).expanduser().resolve()
        rep_path.parent.mkdir(parents=True, exist_ok=True)
        with rep_path.open("w", encoding="utf-8", newline="") as f:
            w = csv.writer(f)
            w.writerow(["email", "ok", "reason"])
            for r in results:
                w.writerow([r["email"], "1" if r["ok"] else "0", r["reason"]])
        print(f"Отчёт: {rep_path}")

    ok_emails = [r["email"] for r in ok]
    if args.dedup:
        ok_emails, removed_dup = dedup_keep_order(ok_emails)
        print(f"Удалено дубликатов: {removed_dup}")

    if args.txt_out:
        txt_path = Path(args.txt_out).expanduser().resolve()
        txt_path.parent.mkdir(parents=True, exist_ok=True)
        txt_path.write_text(", ".join(ok_emails), encoding="utf-8")
        print(f"TXT (через запятую): {txt_path} ({len(ok_emails)} адресов)")

    if args.out:
        out_path = Path(args.out).expanduser().resolve()
        if is_xlsx:
            removed_bad, removed_dup, _removed = _write_xlsx_clean_dedup(
                src, out_path, bad_set, args.dedup, args.email_col, args.start_row,
                drop_bad_syntax=True,
            )
            print(f"Очищенный xlsx: {out_path} (удалено невалидных: {removed_bad}, дубликатов: {removed_dup})")
        else:
            out_path.parent.mkdir(parents=True, exist_ok=True)
            out_path.write_text("\n".join(ok_emails), encoding="utf-8")
            print(f"Очищенный файл: {out_path}")


if __name__ == "__main__":
    main()
