import json
import argparse
import queue
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import cloud_runtime
import email_validator
import email_gui
import send_email


class EmailSyntaxTests(unittest.TestCase):
    def test_rejects_common_garbage(self):
        invalid = [
            "", "   ", ".lead@example.com", "trail.@example.com",
            "two..dots@example.com", "missing-at.example.com", "a@@example.com",
            "user@-example.com", "user@example-.com", "user@example",
            "user@example..com", "user@example.com.", "user @example.com",
        ]
        for value in invalid:
            with self.subTest(value=value):
                self.assertFalse(email_validator.is_valid_email_syntax(value))

        self.assertTrue(email_validator.is_valid_email_syntax("good.user+tag@example.com"))

    def test_sender_does_not_extract_valid_substring_from_broken_address(self):
        self.assertEqual(send_email.extract_emails(".bad@example.com"), [])
        self.assertEqual(send_email.extract_emails("a..b@example.com"), [])
        self.assertEqual(
            send_email.extract_emails("Name <good.user+tag@example.com>"),
            ["good.user+tag@example.com"],
        )

    def test_nxdomain_does_not_use_dns_server_address_as_domain_address(self):
        nxdomain = (
            "Server:  192.0.2.53\nAddress: 192.0.2.53#53\n\n"
            "** server can't find definitely-invalid.example: NXDOMAIN\n"
        )
        with patch.object(email_validator, "_nslookup", return_value=(nxdomain, True, False)):
            email_validator._dns_cache.clear()
            self.assertEqual(
                email_validator.domain_mail_status("definitely-invalid.example"),
                (False, "no_mx"),
            )


class SmtpFailureTests(unittest.TestCase):
    def test_auth_failure_during_retry_writes_failed_progress_and_keeps_cursor(self):
        class FakeSmtp:
            created = 0

            def __init__(self, *_args, **_kwargs):
                type(self).created += 1
                self.number = type(self).created

            def login(self, _user, _password):
                if self.number >= 2:
                    raise send_email.smtplib.SMTPAuthenticationError(535, b"bad auth")

            def send_message(self, _message):
                raise send_email.smtplib.SMTPRecipientsRefused(
                    {"recipient@example.com": (550, b"daily limit")}
                )

            def quit(self):
                return None

        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            template = base / "template.html"
            state = base / "state.json"
            progress = base / "progress.json"
            template.write_text(
                "<html><head><title>Test</title></head><body>Hello</body></html>",
                encoding="utf-8",
            )
            args = argparse.Namespace(
                subject=None,
                smtp_account=[], smtp_host="smtp.example.com", smtp_port=465,
                smtp_user="sender@example.com", smtp_password="secret",
                from_email="sender@example.com", limit_per_day=10,
                limit_per_minute=None, state_file=str(state),
                progress_file=str(progress), campaign_key="smtp-failure-test",
                campaign_id="", dry_run=False,
                hub_url="", hub_connection_id=0, hub_module_secret="",
                hub_insecure_ssl=False, xlsx_kind_filter="ALL",
            )
            recipients = [
                send_email.RecipientRow("recipient@example.com", {}, source_row=2)
            ]
            with patch.object(send_email.smtplib, "SMTP_SSL", FakeSmtp), patch.object(
                send_email.time, "sleep", return_value=None
            ):
                with self.assertRaisesRegex(RuntimeError, "SMTP"):
                    send_email.send_all(args, recipients, template)

            payload = json.loads(progress.read_text(encoding="utf-8"))
            self.assertEqual(payload["status"], "failed")
            self.assertEqual(payload["sent"], 0)
            self.assertFalse(state.exists())


class WorkbookCleaningTests(unittest.TestCase):
    def test_empty_and_invalid_recipient_rows_are_removed(self):
        from openpyxl import Workbook, load_workbook

        with tempfile.TemporaryDirectory() as tmp:
            src = Path(tmp) / "source.xlsx"
            dest = Path(tmp) / "clean.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["name", "email"])
            ws.append(["Good", "good@example.com"])
            ws.append(["No email", None])
            ws.append(["Garbage", "not-an-email"])
            ws.append(["Dots", "a..b@example.com"])
            ws.append(["Duplicate", "GOOD@example.com"])
            wb.save(src)

            emails, rows = email_validator._read_xlsx_emails(src, "B", 2)
            self.assertEqual(rows, [2, 3, 4, 5, 6])
            self.assertEqual(emails[1], "")

            removed_bad, removed_dup, _ = email_validator._write_xlsx_clean_dedup(
                src, dest, set(), True, "B", 2, drop_bad_syntax=True
            )
            self.assertEqual((removed_bad, removed_dup), (3, 1))
            cleaned = load_workbook(dest, data_only=True)
            values = list(cleaned.active.iter_rows(min_row=2, values_only=True))
            cleaned.close()
            self.assertEqual(values, [("Good", "good@example.com")])

    def test_all_sheet_mode_keeps_valid_addresses_from_each_sheet(self):
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "multiple.xlsx"
            wb = Workbook()
            first = wb.active
            first.append(["email"])
            first.append(["one@example.com"])
            second = wb.create_sheet("Second")
            second.append(["email"])
            second.append(["two@example.com"])
            second.append(["a..broken@example.com"])
            wb.save(path)
            recipients = send_email.load_recipients_from_file(
                path, "ALL", "A", [], None, None, 2, False
            )
            self.assertEqual(
                [row.email for row in recipients],
                ["one@example.com", "two@example.com"],
            )


class CloudStateMergeTests(unittest.TestCase):
    def test_merge_never_moves_counters_or_campaign_backwards(self):
        local = {
            "date": "2026-07-10", "sent_today": 3,
            "account_sent_today": {"a": 3},
            "campaigns": {"campaign": {"cursor_index": 5, "last_row": 8}},
        }
        remote = {
            "date": "2026-07-10", "sent_today": 7,
            "account_sent_today": {"a": 6, "b": 1},
            "campaigns": {
                "campaign": {"cursor_index": 9, "last_row": 12},
                "other": {"cursor_index": 2, "last_row": 3},
            },
        }
        merged = cloud_runtime.merge_sending_states(local, remote)
        self.assertEqual(merged["sent_today"], 7)
        self.assertEqual(merged["account_sent_today"], {"a": 6, "b": 1})
        self.assertEqual(merged["campaigns"]["campaign"]["cursor_index"], 9)
        self.assertEqual(merged["campaigns"]["campaign"]["last_row"], 12)
        self.assertIn("other", merged["campaigns"])

    def test_atomic_sender_state_is_valid_json(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "state.json"
            state = send_email.SendingState(path, "campaign")
            state.sent_today = 1
            state.advance_cursor(2, 4)
            payload = json.loads(path.read_text(encoding="utf-8"))
            self.assertEqual(payload["sent_today"], 1)
            self.assertEqual(payload["campaigns"]["campaign"]["cursor_index"], 2)
            self.assertFalse(path.with_suffix(".json.tmp").exists())

    def test_explicit_local_reset_wins_over_larger_remote_cursor(self):
        local = {
            "date": "2026-07-10", "sent_today": 0, "account_sent_today": {},
            "daily_override_at": "2026-07-10T12:00:00",
            "campaigns": {
                "campaign": {
                    "cursor_index": 0, "last_row": 0,
                    "override_at": "2026-07-10T12:00:00",
                }
            },
        }
        remote = {
            "date": "2026-07-10", "sent_today": 50,
            "account_sent_today": {"a": 50},
            "campaigns": {"campaign": {"cursor_index": 50, "last_row": 55}},
        }
        merged = cloud_runtime.merge_sending_states(local, remote)
        self.assertEqual(merged["sent_today"], 0)
        self.assertEqual(merged["account_sent_today"], {})
        self.assertEqual(merged["campaigns"]["campaign"]["cursor_index"], 0)


class CloudCommandTests(unittest.TestCase):
    def test_project_upload_omits_credentials_states_and_unrelated_bundles(self):
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            (base / "send_email.py").write_text("pass", encoding="utf-8")
            (base / "mailer_profiles.json").write_text("secret", encoding="utf-8")
            (base / ".send_email_state.json").write_text("{}", encoding="utf-8")
            (base / "module&hub").mkdir()
            (base / "module&hub" / "large.php").write_text("x", encoding="utf-8")
            names = {
                path.relative_to(base).as_posix()
                for path in cloud_runtime.iter_project_files(base)
            }
            self.assertEqual(names, {"send_email.py"})

    def test_cloud_uses_local_campaign_key_and_syncs_state_before_start(self):
        class FakeRuntime:
            def __init__(self):
                self.synced = []

            def upload_file(self, local_path, remote_path):
                return None

            def sync_state_file(self, local_path, remote_path):
                self.synced.append((Path(local_path), remote_path))
                return True

            def migrate_campaign_key(self, local_path, remote_path, target_key, legacy_key):
                return False

        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            template = base / "mail.html"
            recipients = base / "base.xlsx"
            state = base / "state.json"
            template.write_text("<html></html>", encoding="utf-8")
            recipients.write_bytes(b"xlsx-placeholder")
            state.write_text('{"campaigns": {}}', encoding="utf-8")
            command = [
                "python3", str(email_gui.SCRIPT_PATH),
                "--template", str(template),
                "--xlsx-sheet", "ALL",
                "--xlsx-email-col", "G",
                "--xlsx-start-row", "2",
                "--xlsx-kind-filter", "ALL",
                "--to-file", str(recipients),
                "--state-file", str(state),
            ]
            app = object.__new__(email_gui.MailerApp)
            app.log_queue = queue.Queue()
            runtime = FakeRuntime()
            remote = app._build_remote_command(command, "/srv/mailer", runtime=runtime)

            remote_key = remote[remote.index("--campaign-key") + 1]
            args = argparse.Namespace(
                to_file=str(recipients), xlsx_sheet="ALL", xlsx_email_col="G",
                xlsx_kind_col="", xlsx_kind_filter="ALL", xlsx_start_row=2,
                xlsx_fields="", allow_duplicate_emails=False,
            )
            self.assertEqual(remote_key, send_email.build_campaign_key(args, template))
            self.assertEqual(len(runtime.synced), 1)
            self.assertEqual(runtime.synced[0][0], state.resolve())


if __name__ == "__main__":
    unittest.main()
