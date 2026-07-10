#!/usr/bin/env python3
from __future__ import annotations
import argparse
import html as html_lib
import hashlib
import hmac
import json
import mimetypes
import os
import re
import signal
import smtplib
import ssl
import time
import uuid
from collections import deque
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from email.message import Message
from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import formatdate
from pathlib import Path
from typing import Iterable
from urllib.parse import urlencode
from urllib.request import Request, urlopen

from email_validator import is_valid_email_syntax

_EMAIL_TOKEN_CHARS = r"A-Za-z0-9.!#$%&'*+/=?^_`{|}~+\-"
EMAIL_RE = re.compile(
    rf"(?<![{_EMAIL_TOKEN_CHARS}])"
    rf"[{_EMAIL_TOKEN_CHARS}]+@[A-Za-z0-9.\-]+\.[A-Za-z]{{2,63}}"
    rf"(?![{_EMAIL_TOKEN_CHARS}@])"
)
TITLE_RE = re.compile(r"<title>(.*?)</title>", re.IGNORECASE | re.DOTALL)
TAG_RE = re.compile(r"<[^>]+>")
PLACEHOLDER_RE = re.compile(r"\{\{\s*([A-Za-z][A-Za-z0-9_]*)\s*\}\}")
IMG_SRC_RE = re.compile(r'(<img[^>]+src=["\'])([^"\']+)(["\'])', re.IGNORECASE)
TYPE_TEMPLATE_RULES = [
    (("гвс", "отоплен"), Path("ГВС и отопление") / "landing_gvs_otoplenie.html"),
    (("машиностро",), Path("Машиностроение") / "landing_factories_cooling.html"),
    (("море", "морск"), Path("Морские теплообменники") / "landing_food_industry_unsubscribe_2_updated.html"),
    (("нефтегаз", "нефте", "газ"), Path("Нефтегаз") / "landing_oil_gas.html"),
    (("пищев",), Path("Пищевые компании") / "landing_food_industry_unsubscribe_updated.html"),
    (("проект",), Path("Проектные организации") / "landing_project_orgs.html"),
    (("сахар",), Path("Сахарные заводы") / "landing_sugar_plants.html"),
    (("торгов",), Path("Торговые компании") / "landing_trading_companies.html"),
]
BTN_INLINE_STYLE = (
    "display:inline-block;background:#0077cc;color:#ffffff;text-decoration:none;"
    "padding:12px 20px;border-radius:6px;font-weight:700;margin:8px 6px;"
)
BTN_CONTAINER_INLINE_STYLE = "text-align:center;margin-top:20px;"
CLASS_INLINE_STYLES = {
    "container": (
        "max-width:700px;margin:20px auto;background:#ffffff;padding:24px;border-radius:10px;"
        "box-shadow:0 0 10px rgba(0,0,0,.1);font-family:Arial,sans-serif;line-height:1.6;"
        "color:#000000;"
    ),
    "img-row": "text-align:center;margin-top:15px;font-size:0;",
    "btn-container": BTN_CONTAINER_INLINE_STYLE,
    "signature": "margin-top:30px;font-weight:700;",
    "small": "font-size:12px;color:#777777;text-align:center;",
}
TAG_INLINE_STYLES = {
    "body": "margin:0;padding:0;background:#f9f9f9;",
    "h2": "color:#000000;margin-top:0;",
    "h3": "color:#000000;margin-top:0;",
    "ul": "padding-left:18px;",
    "hr": "border:none;border-top:1px solid #dddddd;margin:30px 0;",
}

STOP_REQUESTED = False
BUILD_MARKER = "cloud-sync-test-20260312-1"


def request_stop(signum, _frame) -> None:
    global STOP_REQUESTED
    STOP_REQUESTED = True
    print(f"Получен сигнал остановки: {signum}. Завершаю текущую задачу корректно...")


@dataclass
class RecipientRow:
    email: str
    fields: dict[str, str]
    source_row: int | None = None


@dataclass
class SmtpAccount:
    host: str
    port: int
    user: str
    password: str
    from_email: str
    daily_limit: int | None = None
    label: str = ""

    @property
    def key(self) -> str:
        return (self.user or self.from_email or self.label).strip().lower()

    @property
    def display_name(self) -> str:
        return self.label or self.from_email or self.user or self.host


class SendingState:
    def __init__(self, state_path: Path, campaign_key: str):
        self.state_path = state_path
        self.campaign_key = campaign_key
        self.current_date = date.today().isoformat()
        self.sent_today = 0
        self.account_sent_today: dict[str, int] = {}
        self.cursor_index = 0
        self.last_row = 0
        self.raw_data: dict = {}
        self._load()

    def _load(self) -> None:
        if not self.state_path.exists():
            self.raw_data = {"date": self.current_date, "sent_today": 0, "campaigns": {}}
            return
        try:
            data = json.loads(self.state_path.read_text(encoding="utf-8"))
        except Exception:
            self.raw_data = {"date": self.current_date, "sent_today": 0, "campaigns": {}}
            return

        if not isinstance(data, dict):
            data = {}
        self.raw_data = data

        previous_date = str(data.get("date", ""))
        if previous_date == self.current_date:
            self.sent_today = int(data.get("sent_today", 0))
            raw_accounts = data.get("account_sent_today", {})
            if isinstance(raw_accounts, dict):
                self.account_sent_today = {
                    str(key): int(value)
                    for key, value in raw_accounts.items()
                    if str(key).strip()
                }
        else:
            self.sent_today = 0
            self.account_sent_today = {}

        campaigns = data.get("campaigns", {})
        if isinstance(campaigns, dict):
            campaign = campaigns.get(self.campaign_key, {})
            if isinstance(campaign, dict):
                self.cursor_index = int(campaign.get("cursor_index", 0))
                self.last_row = int(campaign.get("last_row", 0))
        if "campaigns" not in self.raw_data or not isinstance(self.raw_data.get("campaigns"), dict):
            self.raw_data["campaigns"] = {}

    def save(self) -> None:
        self.raw_data["date"] = self.current_date
        self.raw_data["sent_today"] = self.sent_today
        self.raw_data["account_sent_today"] = self.account_sent_today
        campaigns = self.raw_data.setdefault("campaigns", {})
        existing_campaign = campaigns.get(self.campaign_key, {})
        if not isinstance(existing_campaign, dict):
            existing_campaign = {}
        campaigns[self.campaign_key] = {
            **existing_campaign,
            "cursor_index": self.cursor_index,
            "last_row": self.last_row,
        }
        self.state_path.parent.mkdir(parents=True, exist_ok=True)
        payload = json.dumps(self.raw_data, ensure_ascii=False, indent=2)
        tmp_path = self.state_path.with_suffix(self.state_path.suffix + ".tmp")
        tmp_path.write_text(
            payload,
            encoding="utf-8",
        )
        tmp_path.replace(self.state_path)

    def advance_cursor(self, cursor_index: int, source_row: int | None) -> None:
        self.cursor_index = max(cursor_index, 0)
        if source_row is not None:
            self.last_row = max(source_row, 0)
        self.save()

    def mark_account_sent(self, account: SmtpAccount) -> None:
        key = account.key
        if key:
            self.account_sent_today[key] = self.account_sent_today.get(key, 0) + 1
        self.sent_today += 1


class ProgressReporter:
    def __init__(self, path: str | None):
        self.path = Path(path).expanduser().resolve() if path else None

    def write(self, **payload: object) -> None:
        if not self.path:
            return
        data = {
            "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            **payload,
        }
        try:
            self.path.parent.mkdir(parents=True, exist_ok=True)
            tmp_path = self.path.with_suffix(self.path.suffix + ".tmp")
            tmp_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
            tmp_path.replace(self.path)
        except Exception as error:
            print(f"Не удалось записать progress-файл: {error}")


class RateLimiter:
    def __init__(self, limit_per_minute: int | None):
        self.limit_per_minute = limit_per_minute
        self.sent_timestamps: deque[float] = deque()

    def wait_for_slot(self) -> None:
        if not self.limit_per_minute or self.limit_per_minute <= 0:
            return

        now = time.time()
        while self.sent_timestamps and now - self.sent_timestamps[0] >= 60:
            self.sent_timestamps.popleft()

        if len(self.sent_timestamps) >= self.limit_per_minute:
            wait_seconds = 60 - (now - self.sent_timestamps[0])
            if wait_seconds > 0:
                print(f"Лимит/мин достигнут, жду {wait_seconds:.1f} сек...")
                time.sleep(wait_seconds)

            now = time.time()
            while self.sent_timestamps and now - self.sent_timestamps[0] >= 60:
                self.sent_timestamps.popleft()

    def mark_sent(self) -> None:
        self.sent_timestamps.append(time.time())


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Отправка HTML-писем по шаблону через SMTP (с Excel-полями и лимитами)."
    )
    parser.add_argument(
        "--template",
        required=False,
        default="",
        help="Путь к HTML-шаблону (например: './Нефтегаз/landing_oil_gas.html').",
    )
    parser.add_argument(
        "--subject",
        help="Тема письма. По умолчанию берется из <title>. Можно плейсхолдеры {{A}}, {{B}}.",
    )
    parser.add_argument(
        "--to",
        action="append",
        default=[],
        help="Email получателя. Можно указывать несколько раз.",
    )
    parser.add_argument(
        "--to-file",
        help="Файл получателей: .txt/.csv или .xlsx.",
    )
    parser.add_argument(
        "--xlsx-sheet",
        default="active",
        help="Лист .xlsx: имя листа или номер (1,2,3...). По умолчанию active.",
    )
    parser.add_argument(
        "--xlsx-email-col",
        default="A",
        help="Колонка email в .xlsx (например: A). По умолчанию A.",
    )
    parser.add_argument(
        "--xlsx-fields",
        default="",
        help="Колонки для подстановки в шаблон, через запятую (например: A,B,C,D).",
    )
    parser.add_argument(
        "--xlsx-kind-col",
        default="",
        help="Колонка 'вид рассылки' (например: P). Если задана — шаблон выбирается по значению.",
    )
    parser.add_argument(
        "--xlsx-kind-filter",
        default="ALL",
        help=(
            "Фильтр по виду рассылки из xlsx-kind-col: ALL или список через запятую "
            "(например: 'ГВС' или 'ГВС,Море')."
        ),
    )
    parser.add_argument(
        "--xlsx-start-row",
        type=int,
        default=2,
        help="С какой строки читать .xlsx (по умолчанию 2, чтобы пропустить заголовок).",
    )
    parser.add_argument(
        "--allow-duplicate-emails",
        action="store_true",
        help="Не удалять дубликаты email (полезно для тестов на один адрес).",
    )
    parser.add_argument(
        "--limit-per-minute",
        type=int,
        default=None,
        help="Лимит писем в минуту (например: 20).",
    )
    parser.add_argument(
        "--limit-per-day",
        type=int,
        default=None,
        help="Лимит писем в сутки (например: 300).",
    )
    parser.add_argument(
        "--state-file",
        default=".send_email_state.json",
        help="Файл состояния для суточного лимита.",
    )
    parser.add_argument(
        "--campaign-key",
        default="",
        help="Стабильный ключ кампании (используется GUI для local/cloud state).",
    )
    parser.add_argument(
        "--progress-file",
        default="",
        help="JSON-файл статуса выполнения для GUI/облака.",
    )
    parser.add_argument(
        "--hub-url",
        default="",
        help="Базовый URL Mailroute Hub для трекинг-пикселя (например: https://hub.example.com).",
    )
    parser.add_argument(
        "--hub-connection-id",
        type=int,
        default=0,
        help="ID connection в Hub для трекинг-пикселя.",
    )
    parser.add_argument(
        "--hub-module-secret",
        default="",
        help="module_secret из Hub для подписи URL пикселя.",
    )
    parser.add_argument(
        "--hub-insecure-ssl",
        action="store_true",
        help="Отключить проверку TLS сертификата при запросах к Hub (только для теста).",
    )
    parser.add_argument(
        "--campaign-id",
        default="",
        help="Внешний ID рассылки (если не задан, генерируется автоматически).",
    )
    parser.add_argument(
        "--smtp-host",
        default=os.getenv("SMTP_HOST", "smtp.timeweb.ru"),
        help="SMTP-сервер (по умолчанию: smtp.timeweb.ru).",
    )
    parser.add_argument(
        "--smtp-port",
        type=int,
        default=int(os.getenv("SMTP_PORT", "465")),
        help="SMTP-порт (по умолчанию: 465 для SSL).",
    )
    parser.add_argument(
        "--smtp-user",
        default=os.getenv("SMTP_USER"),
        help="SMTP логин. Можно передать через SMTP_USER.",
    )
    parser.add_argument(
        "--smtp-password",
        default=os.getenv("SMTP_PASSWORD"),
        help="SMTP пароль. Можно передать через SMTP_PASSWORD.",
    )
    parser.add_argument(
        "--from-email",
        dest="from_email",
        help="Email отправителя (по умолчанию = smtp-user).",
    )
    parser.add_argument(
        "--smtp-account",
        action="append",
        default=[],
        help=(
            "Дополнительный SMTP аккаунт в JSON: "
            '{"host":"smtp.timeweb.ru","port":465,"user":"...","password":"...",'
            '"from_email":"...","daily_limit":2000,"label":"domain.ru"}'
        ),
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Проверить сборку писем без реальной отправки.",
    )
    return parser.parse_args()


def normalize_col_name(col: str) -> str:
    name = col.strip().upper()
    if not re.fullmatch(r"[A-Z]+", name):
        raise RuntimeError(f"Некорректная колонка: {col}")
    return name


def parse_columns(raw: str) -> list[str]:
    if not raw.strip():
        return []
    return [normalize_col_name(chunk) for chunk in raw.split(",") if chunk.strip()]


def parse_smtp_accounts(args: argparse.Namespace) -> list[SmtpAccount]:
    accounts: list[SmtpAccount] = []
    if args.smtp_account:
        for raw in args.smtp_account[:5]:
            try:
                payload = json.loads(raw)
            except Exception as error:
                raise RuntimeError(f"Некорректный --smtp-account JSON: {error}") from error
            if not isinstance(payload, dict):
                raise RuntimeError("--smtp-account должен быть JSON-объектом.")
            user = str(payload.get("user") or payload.get("smtp_user") or "").strip()
            password = str(payload.get("password") or payload.get("smtp_password") or "").strip()
            from_email = str(payload.get("from_email") or payload.get("from") or user).strip()
            host = str(payload.get("host") or payload.get("smtp_host") or args.smtp_host).strip()
            port = int(payload.get("port") or payload.get("smtp_port") or args.smtp_port)
            daily_limit_raw = payload.get("daily_limit", args.limit_per_day)
            daily_limit = int(daily_limit_raw) if daily_limit_raw not in (None, "") else None
            label = str(payload.get("label") or "").strip()
            if not user or not password:
                raise RuntimeError("--smtp-account: заполните user и password.")
            accounts.append(
                SmtpAccount(
                    host=host,
                    port=port,
                    user=user,
                    password=password,
                    from_email=from_email or user,
                    daily_limit=daily_limit,
                    label=label,
                )
            )
    else:
        user = args.smtp_user
        password = args.smtp_password
        from_email = args.from_email or user or "no-reply@example.com"
        accounts.append(
            SmtpAccount(
                host=args.smtp_host,
                port=args.smtp_port,
                user=user or "",
                password=password or "",
                from_email=from_email,
                daily_limit=args.limit_per_day,
                label=from_email,
            )
        )
    return accounts


def extract_emails(text: str) -> list[str]:
    if not text:
        return []
    return [candidate for candidate in EMAIL_RE.findall(text) if is_valid_email_syntax(candidate)]


def parse_kind_filter(raw: str) -> set[str] | None:
    value = (raw or "").strip()
    if not value or value.upper() == "ALL":
        return None
    parsed = {chunk.strip().lower() for chunk in value.split(",") if chunk.strip()}
    return parsed or None


def col_to_index(col: str) -> int:
    index = 0
    for char in col:
        index = index * 26 + (ord(char) - ord("A") + 1)
    return index


def extract_subject(html: str, fallback: str) -> str:
    match = TITLE_RE.search(html)
    if not match:
        return fallback
    title = re.sub(r"\s+", " ", match.group(1)).strip()
    return title or fallback


def html_to_text(html: str) -> str:
    text = TAG_RE.sub(" ", html)
    text = re.sub(r"\s+", " ", text).strip()
    return text or "Письмо в HTML-формате."


def to_str(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def select_sheet(workbook, sheet_arg: str):
    if sheet_arg == "active":
        return workbook.active
    if sheet_arg.isdigit():
        index = int(sheet_arg) - 1
        if index < 0 or index >= len(workbook.worksheets):
            raise RuntimeError(f"Лист с номером {sheet_arg} не найден.")
        return workbook.worksheets[index]
    if sheet_arg not in workbook.sheetnames:
        raise RuntimeError(f"Лист '{sheet_arg}' не найден.")
    return workbook[sheet_arg]


def select_sheets(workbook, sheet_arg: str) -> list:
    """Return all worksheets for the GUI's ``ALL`` mode."""
    if (sheet_arg or "").strip().upper() == "ALL":
        return list(workbook.worksheets)
    return [select_sheet(workbook, sheet_arg)]


def load_recipients_from_file(
    path: Path,
    xlsx_sheet: str,
    xlsx_email_col: str,
    xlsx_fields: list[str],
    xlsx_kind_col: str | None,
    xlsx_kind_filter: set[str] | None,
    xlsx_start_row: int,
    allow_duplicate_emails: bool,
) -> list[RecipientRow]:
    ext = path.suffix.lower()
    if ext in {".txt", ".csv"}:
        content = path.read_text(encoding="utf-8", errors="ignore")
        emails = extract_emails(content)
        if not allow_duplicate_emails:
            seen: set[str] = set()
            unique_emails: list[str] = []
            for email in emails:
                key = email.lower()
                if key in seen:
                    continue
                seen.add(key)
                unique_emails.append(email)
            emails = unique_emails
        return [RecipientRow(email=email, fields={}, source_row=None) for email in emails]

    if ext != ".xlsx":
        raise RuntimeError("Поддерживаются только .txt/.csv/.xlsx для --to-file.")

    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise RuntimeError(
            "Для чтения .xlsx установите openpyxl: pip install openpyxl"
        ) from error

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheets = select_sheets(wb, xlsx_sheet)
        email_idx = col_to_index(xlsx_email_col)
        field_indexes = {col: col_to_index(col) for col in xlsx_fields}
        kind_idx = col_to_index(xlsx_kind_col) if xlsx_kind_col else None

        recipients: list[RecipientRow] = []
        seen = set()
        for sheet in sheets:
            for row_idx, row in enumerate(
                sheet.iter_rows(min_row=xlsx_start_row, values_only=True), start=xlsx_start_row
            ):
                email_cell = row[email_idx - 1] if len(row) >= email_idx else None
                emails = extract_emails(to_str(email_cell))
                if not emails:
                    continue

                fields: dict[str, str] = {}
                for col, idx in field_indexes.items():
                    value = row[idx - 1] if len(row) >= idx else None
                    fields[col] = to_str(value)
                    fields[f"COL_{col}"] = to_str(value)
                if kind_idx:
                    kind_value = to_str(row[kind_idx - 1] if len(row) >= kind_idx else None)
                    if xlsx_kind_filter is not None and kind_value.lower() not in xlsx_kind_filter:
                        continue
                    fields["KIND"] = kind_value
                    if xlsx_kind_col:
                        fields[xlsx_kind_col] = kind_value
                        fields[f"COL_{xlsx_kind_col}"] = kind_value
                fields["ROW"] = str(row_idx)
                fields["SHEET"] = to_str(getattr(sheet, "title", ""))

                for email in emails:
                    if not allow_duplicate_emails:
                        key = email.lower()
                        if key in seen:
                            continue
                        seen.add(key)
                    recipients.append(RecipientRow(email=email, fields=dict(fields), source_row=row_idx))
        return recipients
    finally:
        wb.close()


def detect_xlsx_email_columns(path: Path, xlsx_sheet: str, sample_rows: int = 2000) -> list[str]:
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        return []

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheets = select_sheets(wb, xlsx_sheet)
        hits: dict[int, int] = {}
        for sheet in sheets:
            for row in sheet.iter_rows(min_row=1, max_row=sample_rows, values_only=True):
                for col_idx, value in enumerate(row, start=1):
                    text = to_str(value)
                    if text and EMAIL_RE.fullmatch(text) and is_valid_email_syntax(text):
                        hits[col_idx] = hits.get(col_idx, 0) + 1
        sorted_hits = sorted(hits.items(), key=lambda item: item[1], reverse=True)
        return [get_column_letter(col_idx) for col_idx, _ in sorted_hits[:5]]
    finally:
        wb.close()


def iter_inline_images(html: str, template_dir: Path) -> Iterable[tuple[str, Path]]:
    for match in IMG_SRC_RE.finditer(html):
        src = match.group(2).strip()
        lower_src = src.lower()
        if lower_src.startswith(("http://", "https://", "cid:", "data:")):
            continue
        local_path = (template_dir / src).resolve()
        if local_path.exists() and local_path.is_file():
            yield src, local_path


def replace_placeholders(text: str, fields: dict[str, str]) -> str:
    def repl(match: re.Match) -> str:
        key = match.group(1)
        return fields.get(key, match.group(0))

    return PLACEHOLDER_RE.sub(repl, text)


def add_inline_style(tag_html: str, style_chunk: str) -> str:
    style_attr_re = re.compile(r'style=["\']([^"\']*)["\']', re.IGNORECASE)
    match = style_attr_re.search(tag_html)
    if match:
        current = match.group(1).strip()
        new_style = f"{current};{style_chunk}" if current else style_chunk
        return style_attr_re.sub(f'style="{new_style}"', tag_html, count=1)
    return tag_html[:-1] + f' style="{style_chunk}">'


def make_email_friendly_html(html: str) -> str:
    for class_name, class_style in CLASS_INLINE_STYLES.items():
        html = re.sub(
            rf'(<[a-z0-9]+\b[^>]*class=["\'][^"\']*\b{re.escape(class_name)}\b[^"\']*["\'][^>]*>)',
            lambda m, style=class_style: add_inline_style(m.group(1), style),
            html,
            flags=re.IGNORECASE,
        )

    for tag_name, tag_style in TAG_INLINE_STYLES.items():
        html = re.sub(
            rf"(<{tag_name}\b[^>]*>)",
            lambda m, style=tag_style: add_inline_style(m.group(1), style),
            html,
            flags=re.IGNORECASE,
        )

    def style_btn_open_tag(match: re.Match) -> str:
        return add_inline_style(match.group(1), BTN_INLINE_STYLE)

    def style_btn_container_tag(match: re.Match) -> str:
        return add_inline_style(match.group(1), BTN_CONTAINER_INLINE_STYLE)

    html = re.sub(
        r'(<a\b[^>]*class=["\'][^"\']*\bbtn\b[^"\']*["\'][^>]*>)',
        style_btn_open_tag,
        html,
        flags=re.IGNORECASE,
    )
    html = re.sub(
        r'(<div\b[^>]*class=["\'][^"\']*\bbtn-container\b[^"\']*["\'][^>]*>)',
        style_btn_container_tag,
        html,
        flags=re.IGNORECASE,
    )

    def style_images_inside_img_row(match: re.Match) -> str:
        block = match.group(0)
        return re.sub(
            r'(<img\b[^>]*>)',
            lambda img_m: add_inline_style(
                img_m.group(1),
                "width:30%;max-width:220px;min-width:140px;height:auto;"
                "display:inline-block;vertical-align:top;margin:4px;border-radius:8px;object-fit:cover;",
            ),
            block,
            flags=re.IGNORECASE,
        )

    html = re.sub(
        r'<div\b[^>]*class=["\'][^"\']*\bimg-row\b[^"\']*["\'][^>]*>.*?</div>',
        style_images_inside_img_row,
        html,
        flags=re.IGNORECASE | re.DOTALL,
    )

    def anchor_to_button(match: re.Match) -> str:
        full_anchor = match.group(0)
        attrs = match.group(1)
        label = match.group(2).strip()
        href_match = re.search(r'href=["\']([^"\']+)["\']', attrs, re.IGNORECASE)
        href = href_match.group(1) if href_match else "#"
        return (
            '<table role="presentation" border="0" cellpadding="0" cellspacing="0" '
            'style="display:inline-table;margin:8px 6px;">'
            '<tr><td align="center" bgcolor="#0077cc" style="border-radius:6px;">'
            f'<a href="{href}" target="_blank" '
            'style="display:inline-block;padding:12px 20px;font-family:Arial,sans-serif;'
            'font-size:20px;line-height:1.2;font-weight:700;color:#ffffff;'
            f'text-decoration:none;border-radius:6px;">{label}</a>'
            "</td></tr></table>"
        )

    html = re.sub(
        r'<a\b([^>]*class=["\'][^"\']*\bbtn\b[^"\']*["\'][^>]*)>(.*?)</a>',
        anchor_to_button,
        html,
        flags=re.IGNORECASE | re.DOTALL,
    )
    return html


def extract_unsubscribe_url(html: str) -> str:
    hrefs = re.findall(r'<a\b[^>]*href=["\']([^"\']+)["\']', html, flags=re.IGNORECASE)
    for href in hrefs:
        lower = href.lower()
        if "unsubscribe" in lower or "отпис" in lower or "spasibo" in lower:
            return href
    return ""


def build_message(
    from_email: str,
    recipient_email: str,
    subject: str,
    html: str,
    template_dir: Path,
    message_id: str | None = None,
) -> Message:
    msg_root = MIMEMultipart("related")
    msg_root["From"] = from_email
    msg_root["To"] = recipient_email
    msg_root["Subject"] = subject
    msg_root["Date"] = formatdate(localtime=True)
    message_id_domain = from_email.split("@", 1)[1] if "@" in from_email else "localhost"
    mid = message_id or uuid.uuid4().hex
    msg_root["Message-ID"] = f"<{mid}@{message_id_domain}>"
    msg_root["Precedence"] = "bulk"

    msg_alt = MIMEMultipart("alternative")
    msg_root.attach(msg_alt)
    msg_alt.attach(MIMEText(html_to_text(html), "plain", "utf-8"))

    image_map: dict[str, tuple[str, bytes, str, str]] = {}
    html_with_cid = html

    for src, local_path in iter_inline_images(html, template_dir):
        if src in image_map:
            cid = image_map[src][0]
        else:
            cid = uuid.uuid4().hex
            mime_type, _ = mimetypes.guess_type(local_path.name)
            if mime_type is None:
                mime_type = "application/octet-stream"
            maintype, subtype = mime_type.split("/", 1)
            image_map[src] = (cid, local_path.read_bytes(), maintype, subtype)
        html_with_cid = html_with_cid.replace(f'"{src}"', f'"cid:{cid}"')
        html_with_cid = html_with_cid.replace(f"'{src}'", f"'cid:{cid}'")

    msg_alt.attach(MIMEText(html_with_cid, "html", "utf-8"))

    unsubscribe_url = extract_unsubscribe_url(html_with_cid)
    if unsubscribe_url:
        msg_root["List-Unsubscribe"] = f"<{unsubscribe_url}>"
        msg_root["List-Unsubscribe-Post"] = "List-Unsubscribe=One-Click"

    for src, (cid, content, maintype, subtype) in image_map.items():
        if maintype != "image":
            continue
        image_part = MIMEImage(content, _subtype=subtype)
        image_part.add_header("Content-ID", f"<{cid}>")
        image_part.add_header("Content-Disposition", "inline", filename=Path(src).name)
        image_part.add_header("Content-Location", src)
        image_part.add_header("X-Attachment-Id", cid)
        msg_root.attach(image_part)

    return msg_root


def build_recipients(args: argparse.Namespace) -> list[RecipientRow]:
    recipients: list[RecipientRow] = []
    seen = set()

    for email in args.to:
        for parsed in extract_emails(email):
            if not args.allow_duplicate_emails:
                key = parsed.lower()
                if key in seen:
                    continue
                seen.add(key)
            recipients.append(RecipientRow(email=parsed, fields={}, source_row=None))

    if args.to_file:
        to_file_path = Path(args.to_file).expanduser().resolve()
        email_col = normalize_col_name(args.xlsx_email_col)
        kind_col = normalize_col_name(args.xlsx_kind_col) if args.xlsx_kind_col.strip() else None
        kind_filter = parse_kind_filter(args.xlsx_kind_filter)
        if kind_filter is not None and kind_col is None:
            raise RuntimeError("Для фильтра вида укажите --xlsx-kind-col (например P).")
        from_file = load_recipients_from_file(
            path=to_file_path,
            xlsx_sheet=args.xlsx_sheet,
            xlsx_email_col=email_col,
            xlsx_fields=parse_columns(args.xlsx_fields),
            xlsx_kind_col=kind_col,
            xlsx_kind_filter=kind_filter,
            xlsx_start_row=args.xlsx_start_row,
            allow_duplicate_emails=args.allow_duplicate_emails,
        )
        if not from_file and to_file_path.suffix.lower() == ".xlsx":
            suggestions = detect_xlsx_email_columns(to_file_path, args.xlsx_sheet)
            hint = ""
            if suggestions:
                hint = f" Возможно, email-колонка: {', '.join(suggestions)}."
            raise RuntimeError(
                f"В .xlsx не найдено email в колонке {email_col} начиная со строки "
                f"{args.xlsx_start_row}.{hint}"
            )
        for row in from_file:
            if not args.allow_duplicate_emails:
                key = row.email.lower()
                if key in seen:
                    continue
            recipients.append(row)
            if not args.allow_duplicate_emails:
                seen.add(row.email.lower())

    if not recipients:
        raise RuntimeError("Не указаны получатели. Используйте --to или --to-file.")

    return recipients


def pick_template_by_kind(default_template_path: Path, kind: str) -> Path:
    kind_lower = kind.lower()
    if not kind_lower:
        return default_template_path

    base_dir = default_template_path.parent.parent
    for keywords, relative_path in TYPE_TEMPLATE_RULES:
        if any(keyword in kind_lower for keyword in keywords):
            candidate = (base_dir / relative_path).resolve()
            if candidate.exists():
                return candidate
    return default_template_path


def find_first_available_template(base_dir: Path) -> Path | None:
    for _, relative_path in TYPE_TEMPLATE_RULES:
        candidate = (base_dir / relative_path).resolve()
        if candidate.exists():
            return candidate
    return None


def build_campaign_key(args: argparse.Namespace, template_path: Path) -> str:
    raw = "|".join(
        [
            str(Path(args.to_file).expanduser().resolve()) if args.to_file else "",
            args.xlsx_sheet,
            args.xlsx_email_col,
            args.xlsx_kind_col,
            args.xlsx_kind_filter,
            str(args.xlsx_start_row),
            args.xlsx_fields,
            str(args.allow_duplicate_emails),
            str(template_path.resolve()),
        ]
    )
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:24]


def format_wait_until_midnight() -> str:
    now = datetime.now()
    next_midnight = (now + timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
    remaining = max(int((next_midnight - now).total_seconds()), 0)
    days, rem = divmod(remaining, 86400)
    hours, rem = divmod(rem, 3600)
    minutes, seconds = divmod(rem, 60)
    return f"{days}d {hours}h {minutes}m {seconds}s"


def build_open_pixel_url(
    hub_url: str,
    connection_id: int,
    module_secret: str,
    campaign_id: str,
    message_id: str,
    recipient_email: str,
    mail_kind: str = "",
) -> str:
    recipient_email_norm = recipient_email.strip().lower()
    recipient_hash = hashlib.sha256(recipient_email_norm.encode("utf-8")).hexdigest()
    mail_kind_norm = mail_kind.strip()
    ts = str(int(time.time()))
    payload = f"{connection_id}\n{campaign_id}\n{message_id}\n{recipient_hash}\n{recipient_email_norm}\n{ts}"
    signature = hmac.new(
        module_secret.encode("utf-8"),
        payload.encode("utf-8"),
        hashlib.sha256,
    ).hexdigest()
    query = urlencode(
        {
            "cid": connection_id,
            "cmp": campaign_id,
            "mid": message_id,
            "rh": recipient_hash,
            "re": recipient_email_norm,
            "k": mail_kind_norm,
            "ts": ts,
            "sig": signature,
        }
    )
    return hub_url.rstrip("/") + "/index.php/api/v1/public/mail-open.gif?" + query


def add_tracking_pixel(html: str, pixel_url: str) -> str:
    safe_url = html_lib.escape(pixel_url, quote=True)
    pixel_html = (
        f'<img src="{safe_url}" width="1" height="1" alt="" '
        'style="width:1px;height:1px;opacity:0;border:0;" />'
    )
    if "</body>" in html.lower():
        return re.sub(r"</body>", pixel_html + "</body>", html, flags=re.IGNORECASE, count=1)
    return html + pixel_html


def send_campaign_report_to_hub(
    hub_url: str,
    connection_id: int,
    module_secret: str,
    payload: dict,
    insecure_ssl: bool = False,
) -> None:
    if not (hub_url and connection_id > 0 and module_secret):
        return

    raw = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    ts = str(int(time.time()))
    signed = f"{connection_id}\n{ts}\n{raw}"
    sig = hmac.new(module_secret.encode("utf-8"), signed.encode("utf-8"), hashlib.sha256).hexdigest()
    url = hub_url.rstrip("/") + "/index.php/api/v1/public/campaign-report?" + urlencode(
        {"cid": connection_id, "ts": ts, "sig": sig}
    )
    req = Request(
        url=url,
        method="POST",
        data=raw.encode("utf-8"),
        headers={"Content-Type": "application/json"},
    )
    context = ssl._create_unverified_context() if insecure_ssl else None
    with urlopen(req, timeout=20, context=context) as response:
        response.read()


def send_all(args: argparse.Namespace, recipients: list[RecipientRow], template_path: Path) -> None:
    print(f"Build marker: {BUILD_MARKER}")
    progress = ProgressReporter(args.progress_file)
    templates_cache: dict[Path, tuple[str, str, Path]] = {}

    def get_template_content(current_template_path: Path) -> tuple[str, str, Path]:
        cached = templates_cache.get(current_template_path)
        if cached:
            return cached
        html_template = current_template_path.read_text(encoding="utf-8")
        default_subject = extract_subject(html_template, "Коммерческое предложение")
        payload = (html_template, default_subject, current_template_path.parent)
        templates_cache[current_template_path] = payload
        return payload

    _, initial_subject, _ = get_template_content(template_path)
    subject_template = args.subject or initial_subject

    accounts = parse_smtp_accounts(args)
    if len(accounts) > 5:
        raise RuntimeError("Можно указать не более 5 SMTP аккаунтов.")
    from_email = accounts[0].from_email
    campaign_key = (args.campaign_key or "").strip() or build_campaign_key(args, template_path)
    state = SendingState(Path(args.state_file).expanduser().resolve(), campaign_key=campaign_key)
    limiter = RateLimiter(args.limit_per_minute)

    if args.limit_per_day is not None and args.limit_per_day <= 0:
        raise RuntimeError("--limit-per-day должен быть > 0")
    if args.limit_per_minute is not None and args.limit_per_minute <= 0:
        raise RuntimeError("--limit-per-minute должен быть > 0")
    active_accounts = [
        account
        for account in accounts
        if account.daily_limit is None or state.account_sent_today.get(account.key, 0) < account.daily_limit
    ]
    if not active_accounts:
        wait_text = format_wait_until_midnight()
        raise RuntimeError(f"ratelimit wait {wait_text}")

    run_started_at = datetime.now()
    campaign_id = args.campaign_id.strip() or f"cmp-{run_started_at.strftime('%Y%m%d-%H%M%S')}-{uuid.uuid4().hex[:8]}"
    start_cursor = max(state.cursor_index, 0)
    if start_cursor >= len(recipients):
        print("Все строки для этой базы уже отправлены. Сбросьте state-файл для повтора.")
        progress.write(
            status="completed",
            total=0,
            processed=0,
            sent=0,
            failed=0,
            skipped=0,
            percent=100,
            message="Все строки уже отправлены.",
        )
        return
    recipients = recipients[start_cursor:]
    total_to_process = len(recipients)
    print(f"Всего получателей (осталось): {total_to_process}")
    print("SMTP аккаунты:")
    for account in accounts:
        sent_for_account = state.account_sent_today.get(account.key, 0)
        limit_view = "без лимита" if account.daily_limit is None else str(account.daily_limit)
        print(f"  - {account.display_name}: сегодня {sent_for_account}/{limit_view}")
    progress.write(
        status="starting",
        total=total_to_process,
        processed=0,
        sent=0,
        failed=0,
        skipped=0,
        percent=0,
        template=str(template_path),
        accounts=[account.display_name for account in accounts],
        message="Подготовка отправки.",
    )

    if args.dry_run:
        preview = recipients[:3]
        print("DRY RUN OK")
        print(f"From: {from_email}")
        print(f"Всего получателей (осталось): {len(recipients)}")
        print(f"Текущий счётчик в state-файле за сегодня: {state.sent_today}")
        print(f"Курсор кампании: {start_cursor}")
        print(f"Campaign ID: {campaign_id}")
        if state.last_row:
            print(f"Последняя отправленная строка Excel: {state.last_row}")
        if args.hub_url and args.hub_connection_id > 0 and args.hub_module_secret:
            print(f"Трекинг открытий: ON ({args.hub_url}, connection={args.hub_connection_id})")
        if args.limit_per_day:
            print(f"Лимит в сутки: {args.limit_per_day}")
            available = max(args.limit_per_day - state.sent_today, 0)
            print(f"Доступно к отправке сегодня: {available}")
        if args.limit_per_minute:
            print(f"Лимит в минуту: {args.limit_per_minute}")
        print(f"Template: {template_path}")
        print("Примеры первых писем:")
        for row in preview:
            kind = row.fields.get("KIND", "")
            row_template_path = pick_template_by_kind(template_path, kind) if kind else template_path
            personalized_subject = replace_placeholders(subject_template, row.fields)
            print(
                f"  -> {row.email} | subject: {personalized_subject} | "
                f"template: {row_template_path.name}"
            )
        return

    if any(not account.user or not account.password for account in accounts):
        raise RuntimeError(
            "Нужны SMTP логин и пароль. Передайте --smtp-user/--smtp-password "
            "или заполните SMTP аккаунты."
        )

    sent = 0
    failed = 0
    skipped_daily = 0
    run_status = "completed"
    fatal_error: Exception | None = None
    current_account: SmtpAccount | None = None
    smtp: smtplib.SMTP_SSL | None = None

    def close_smtp() -> None:
        nonlocal smtp
        if smtp is not None:
            try:
                smtp.quit()
            except Exception:
                pass
        smtp = None

    def pick_next_account() -> SmtpAccount | None:
        for account in accounts:
            if account.daily_limit is None:
                return account
            if state.account_sent_today.get(account.key, 0) < account.daily_limit:
                return account
        return None

    def connect_account(account: SmtpAccount) -> smtplib.SMTP_SSL:
        connection = smtplib.SMTP_SSL(account.host, account.port, timeout=120)
        connection.login(account.user, account.password)
        return connection

    try:
        for offset, recipient in enumerate(recipients, start=1):
            processed = offset - 1
            if STOP_REQUESTED:
                run_status = "stopped"
                skipped_daily = len(recipients) - (offset - 1)
                print("Остановка подтверждена. Завершаю рассылку после текущего состояния.")
                break
            account = pick_next_account()
            if account is None:
                skipped_daily = len(recipients) - (offset - 1)
                break
            if current_account is None or current_account.key != account.key:
                close_smtp()
                current_account = account
                print(f"Переключение SMTP аккаунта: {account.display_name}")
                try:
                    smtp = connect_account(account)
                except smtplib.SMTPAuthenticationError as error:
                    fatal_error = error
                    run_status = "failed"
                    print(
                        f"Фатальная ошибка SMTP-авторизации "
                        f"для {account.display_name}: {error}"
                    )
                    break

            limiter.wait_for_slot()
            kind = recipient.fields.get("KIND", "")
            row_template_path = pick_template_by_kind(template_path, kind) if kind else template_path
            if not row_template_path.exists():
                raise RuntimeError(f"HTML-шаблон не найден: {row_template_path}")
            row_html_template, _, row_template_dir = get_template_content(row_template_path)
            personalized_html = replace_placeholders(row_html_template, recipient.fields)
            personalized_html = make_email_friendly_html(personalized_html)
            if not personalized_html.strip():
                raise RuntimeError(f"HTML-шаблон пустой: {row_template_path}")
            message_id = uuid.uuid4().hex
            if args.hub_url and args.hub_connection_id > 0 and args.hub_module_secret:
                pixel_url = build_open_pixel_url(
                    hub_url=args.hub_url,
                    connection_id=args.hub_connection_id,
                    module_secret=args.hub_module_secret,
                    campaign_id=campaign_id,
                    message_id=message_id,
                    recipient_email=recipient.email,
                    mail_kind=kind,
                )
                personalized_html = add_tracking_pixel(personalized_html, pixel_url)
            personalized_subject = replace_placeholders(subject_template, recipient.fields)
            message = build_message(
                from_email=account.from_email,
                recipient_email=recipient.email,
                subject=personalized_subject,
                html=personalized_html,
                template_dir=row_template_dir,
                message_id=message_id,
            )

            last_error: Exception | None = None
            for attempt in range(1, 4):
                try:
                    if smtp is None:
                        smtp = connect_account(account)
                    smtp.send_message(message)
                    last_error = None
                    break
                except (TimeoutError, smtplib.SMTPServerDisconnected, smtplib.SMTPException) as error:
                    last_error = error
                    print(f"Ошибка отправки ({attempt}/3) для {recipient.email}: {error}")
                    try:
                        if smtp is not None:
                            smtp.quit()
                    except Exception:
                        pass
                    smtp = None
                    if isinstance(error, smtplib.SMTPAuthenticationError):
                        fatal_error = error
                        run_status = "failed"
                        break
                    if attempt < 3:
                        time.sleep(2)
                        try:
                            smtp = connect_account(account)
                        except smtplib.SMTPAuthenticationError as auth_error:
                            last_error = auth_error
                            fatal_error = auth_error
                            run_status = "failed"
                            print(
                                f"Фатальная ошибка SMTP-авторизации "
                                f"для {account.display_name}: {auth_error}"
                            )
                            break

            if fatal_error is not None:
                break

            if last_error:
                failed += 1
                print(f"[ERR] Не отправлено: {recipient.email} | причина: {last_error}")
                progress.write(
                    status="running",
                    total=total_to_process,
                    processed=min(processed + 1, total_to_process),
                    sent=sent,
                    failed=failed,
                    skipped=skipped_daily,
                    percent=round((min(processed + 1, total_to_process) / total_to_process) * 100, 2),
                    current_email=recipient.email,
                    current_template=str(row_template_path),
                    current_account=account.display_name,
                    message=f"Ошибка отправки: {recipient.email}",
                )
                continue

            limiter.mark_sent()
            sent += 1
            state.mark_account_sent(account)
            state.advance_cursor(start_cursor + offset, recipient.source_row)
            processed = offset
            percent = round((processed / total_to_process) * 100, 2) if total_to_process else 100
            print(
                f"[{sent}] Отправлено: {recipient.email} | "
                f"шаблон: {row_template_path.name} | аккаунт: {account.display_name} | {percent}%"
            )
            progress.write(
                status="running",
                total=total_to_process,
                processed=processed,
                sent=sent,
                failed=failed,
                skipped=skipped_daily,
                percent=percent,
                current_email=recipient.email,
                current_template=str(row_template_path),
                current_account=account.display_name,
                account_sent_today=state.account_sent_today,
                message=f"Отправлено {sent} из {total_to_process}.",
            )
    finally:
        close_smtp()

    print(f"Готово. Отправлено: {sent}.")
    print(f"Статус выполнения: {run_status}")
    if skipped_daily:
        print(
            f"Пропущено из-за суточного лимита: {skipped_daily}. "
            "Запустите снова завтра или увеличьте --limit-per-day."
        )
    final_processed = min(sent + failed + skipped_daily, total_to_process)
    progress.write(
        status=run_status,
        total=total_to_process,
        processed=final_processed,
        sent=sent,
        failed=failed,
        skipped=skipped_daily,
        percent=round((final_processed / total_to_process) * 100, 2) if total_to_process else 100,
        account_sent_today=state.account_sent_today,
        message=(
            f"Фатальная ошибка SMTP: {fatal_error}"
            if fatal_error is not None
            else f"Готово. Отправлено: {sent}."
        ),
    )

    run_finished_at = datetime.now()
    if args.hub_url and args.hub_connection_id > 0 and args.hub_module_secret:
        report_payload = {
            "event": "mailing_campaign",
            "campaign_id": campaign_id,
            "started_at": run_started_at.strftime("%Y-%m-%d %H:%M:%S"),
            "finished_at": run_finished_at.strftime("%Y-%m-%d %H:%M:%S"),
            "success_count": sent,
            "failed_count": failed,
            "skipped_daily_count": skipped_daily,
            "total_count": total_to_process,
            "status": run_status,
            "from_email": from_email,
            "template": str(template_path),
            "kind_filter": args.xlsx_kind_filter,
            "cursor_start": start_cursor,
            "cursor_end": state.cursor_index,
        }
        try:
            send_campaign_report_to_hub(
                hub_url=args.hub_url,
                connection_id=args.hub_connection_id,
                module_secret=args.hub_module_secret,
                payload=report_payload,
                insecure_ssl=args.hub_insecure_ssl,
            )
            print("Отчёт рассылки отправлен в Hub.")
        except Exception as error:
            print(f"Не удалось отправить отчёт в Hub: {error}")

    if fatal_error is not None:
        raise RuntimeError(f"Фатальная ошибка SMTP: {fatal_error}") from fatal_error


def main() -> None:
    signal.signal(signal.SIGTERM, request_stop)
    signal.signal(signal.SIGINT, request_stop)
    args = parse_args()
    if args.template.strip():
        template_path = Path(args.template).expanduser().resolve()
    else:
        detected = find_first_available_template(Path(__file__).resolve().parent)
        if detected is None:
            raise RuntimeError("Не найден ни один HTML-шаблон рядом со скриптом.")
        template_path = detected
        print(f"Автоподбор шаблона: {template_path}")

    if not template_path.exists():
        raise FileNotFoundError(f"Шаблон не найден: {template_path}")
    recipients = build_recipients(args)
    send_all(args, recipients, template_path)


if __name__ == "__main__":
    main()
